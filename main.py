"""
Sistema de Punto de Venta - API principal
Ejecutar con:  uvicorn main:app --host 0.0.0.0 --port 8000 --reload
"""
import io
import math
import secrets
import httpx
from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from datetime import datetime, date, timedelta

import models
import schemas
import auth
from database import engine, get_db, Base, SessionLocal

# Crea las tablas si no existen
Base.metadata.create_all(bind=engine)


def _migrar_esquema():
    """
    Migración ligera para bases de datos (pos.db) creadas con una versión
    anterior del esquema: Base.metadata.create_all solo crea tablas nuevas
    (como 'proveedores'), pero no agrega columnas nuevas a tablas que ya
    existían (como 'proveedor_id' en 'productos'). Aquí se agrega esa
    columna a mano si hace falta, sin tocar los datos existentes.
    """
    with engine.connect() as conn:
        columnas = [fila[1] for fila in conn.exec_driver_sql("PRAGMA table_info(productos)").fetchall()]
        if "proveedor_id" not in columnas:
            conn.exec_driver_sql("ALTER TABLE productos ADD COLUMN proveedor_id INTEGER REFERENCES proveedores(id)")
            conn.commit()
        if "stock_maximo" not in columnas:
            conn.exec_driver_sql("ALTER TABLE productos ADD COLUMN stock_maximo FLOAT DEFAULT 0")
            # Arranque razonable para productos que ya existían: como nunca
            # se registró un máximo, se toma el mayor entre el stock actual
            # y el mínimo, para no sugerir pedidos absurdos el primer día.
            # De ahí en adelante el máximo sube solo con cada entrada de
            # mercancía que supere ese nivel.
            conn.exec_driver_sql(
                "UPDATE productos SET stock_maximo = "
                "CASE WHEN stock > stock_minimo THEN stock ELSE stock_minimo END"
            )
            conn.commit()

        columnas_mov = [fila[1] for fila in conn.exec_driver_sql("PRAGMA table_info(movimientos_inventario)").fetchall()]
        if "costo_unitario" not in columnas_mov:
            conn.exec_driver_sql("ALTER TABLE movimientos_inventario ADD COLUMN costo_unitario FLOAT")
            conn.commit()


_migrar_esquema()

app = FastAPI(title="POS - Sistema de Punto de Venta")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.exception_handler(RequestValidationError)
async def manejar_errores_de_validacion(request: Request, exc: RequestValidationError):
    # Por defecto FastAPI regresa "detail" como una lista de objetos, pero el
    # frontend (api() en app.js) espera un string simple en data.detail.
    # Aquí lo aplanamos, por ejemplo para mostrar bien el mensaje de
    # "la contraseña no cumple con..." al crear/editar un usuario.
    mensajes = [err.get("msg", "Dato inválido") for err in exc.errors()]
    return JSONResponse(status_code=422, content={"detail": " / ".join(mensajes)})


@app.on_event("startup")
def crear_admin_por_defecto():
    """Si no existe ningún usuario, crea un admin inicial (usuario: admin / clave: admin123)."""
    db = SessionLocal()
    try:
        if db.query(models.Usuario).count() == 0:
            admin = models.Usuario(
                username="admin",
                nombre_completo="Administrador",
                password_hash=auth.hash_password("admin123"),
                rol="admin",
            )
            db.add(admin)
            db.commit()
            print("=" * 60)
            print(" Usuario admin creado por defecto:")
            print("   usuario:  admin")
            print("   clave:    admin123")
            print(" ¡Cámbiala en cuanto entres al sistema!")
            print("=" * 60)
    finally:
        db.close()


# ============================================================
#  AUTENTICACIÓN
# ============================================================

@app.post("/api/auth/login", response_model=schemas.TokenResponse)
def login(datos: schemas.LoginRequest, db: Session = Depends(get_db)):
    usuario = db.query(models.Usuario).filter(models.Usuario.username == datos.username).first()
    if not usuario or not auth.verificar_password(datos.password, usuario.password_hash):
        raise HTTPException(401, "Usuario o contraseña incorrectos")
    if not usuario.activo:
        raise HTTPException(401, "Este usuario está deshabilitado")

    token = auth.crear_sesion(db, usuario)
    return schemas.TokenResponse(token=token, usuario=usuario)


@app.get("/api/auth/me", response_model=schemas.UsuarioOut)
def quien_soy(usuario: models.Usuario = Depends(auth.obtener_usuario_actual)):
    return usuario


@app.post("/api/auth/logout")
def logout(authorization: str = None, db: Session = Depends(get_db),
           usuario: models.Usuario = Depends(auth.obtener_usuario_actual)):
    db.query(models.Sesion).filter(models.Sesion.usuario_id == usuario.id).delete()
    db.commit()
    return {"ok": True}


# ============================================================
#  USUARIOS (solo admin)
# ============================================================

@app.post("/api/usuarios", response_model=schemas.UsuarioOut)
def crear_usuario(
    datos: schemas.UsuarioCrear,
    db: Session = Depends(get_db),
    _admin: models.Usuario = Depends(auth.requiere_rol("admin")),
):
    if datos.rol not in ("admin", "cajero"):
        raise HTTPException(400, "El rol debe ser 'admin' o 'cajero'")
    if db.query(models.Usuario).filter(models.Usuario.username == datos.username).first():
        raise HTTPException(400, "Ese nombre de usuario ya existe")

    nuevo = models.Usuario(
        username=datos.username,
        nombre_completo=datos.nombre_completo,
        password_hash=auth.hash_password(datos.password),
        rol=datos.rol,
    )
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo


@app.get("/api/usuarios", response_model=List[schemas.UsuarioOut])
def listar_usuarios(
    db: Session = Depends(get_db),
    _admin: models.Usuario = Depends(auth.requiere_rol("admin")),
):
    return db.query(models.Usuario).order_by(models.Usuario.username).all()


@app.put("/api/usuarios/{usuario_id}", response_model=schemas.UsuarioOut)
def actualizar_usuario(
    usuario_id: int,
    datos: schemas.UsuarioActualizar,
    db: Session = Depends(get_db),
    admin: models.Usuario = Depends(auth.requiere_rol("admin")),
):
    usuario = db.query(models.Usuario).get(usuario_id)
    if not usuario:
        raise HTTPException(404, "Usuario no encontrado")
    if usuario.id == admin.id and datos.activo is False:
        raise HTTPException(400, "No puedes deshabilitarte a ti mismo")

    cambios = datos.dict(exclude_unset=True)
    if "password" in cambios and cambios["password"]:
        usuario.password_hash = auth.hash_password(cambios.pop("password"))
    else:
        cambios.pop("password", None)
    for campo, valor in cambios.items():
        setattr(usuario, campo, valor)

    db.commit()
    db.refresh(usuario)
    return usuario


# ============================================================
#  PERMISOS POR ROL (Administración > Permisos)
# ============================================================

def _obtener_permisos_cajero(db: Session) -> dict:
    """
    Dict { "productos.ver": bool, "productos.agregar": bool, ... } para el
    rol 'cajero'. Completa con False las claves sin registro.

    Compatibilidad con el esquema anterior (un solo bool por módulo, ej.
    "productos"=True): si existe un registro legacy sin punto, se expande
    a todas las acciones de ese módulo.
    """
    permisos = {clave: False for clave in schemas.CLAVES_PERMISO}
    filas = db.query(models.PermisoRol).filter(models.PermisoRol.rol == "cajero").all()
    for fila in filas:
        if fila.modulo in permisos:
            permisos[fila.modulo] = bool(fila.permitido)
        elif fila.modulo in schemas.MODULOS_LEGACY and fila.permitido:
            # Esquema antiguo: un permiso por módulo → todas las acciones
            for clave in schemas.CLAVES_PERMISO:
                if clave.startswith(fila.modulo + "."):
                    permisos[clave] = True
    return permisos


@app.get("/api/permisos")
def obtener_permisos(
    db: Session = Depends(get_db),
    _admin: models.Usuario = Depends(auth.requiere_rol("admin")),
):
    """Permisos actuales del rol 'cajero'. Pantalla Administración > Permisos."""
    return _obtener_permisos_cajero(db)


@app.put("/api/permisos")
def actualizar_permisos(
    datos: schemas.PermisosActualizar,
    db: Session = Depends(get_db),
    _admin: models.Usuario = Depends(auth.requiere_rol("admin")),
):
    # Pydantic v1: .dict(); v2: .model_dump(). Preferimos dict() con fallback.
    try:
        cambios = datos.dict(exclude_unset=True)
    except Exception:
        cambios = datos.model_dump(exclude_unset=True)
    # Solo aceptamos claves conocidas (evita basura en la BD).
    for clave, permitido in cambios.items():
        if clave not in schemas.CLAVES_PERMISO:
            continue
        fila = (
            db.query(models.PermisoRol)
            .filter(models.PermisoRol.rol == "cajero", models.PermisoRol.modulo == clave)
            .first()
        )
        if fila:
            fila.permitido = bool(permitido)
        else:
            db.add(models.PermisoRol(rol="cajero", modulo=clave, permitido=bool(permitido)))
    db.commit()
    return _obtener_permisos_cajero(db)


@app.get("/api/permisos/mias")
def mis_permisos(
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(auth.obtener_usuario_actual),
):
    """Permisos efectivos del usuario que llama (para que el frontend
    sepa qué mostrar en el menú). Admin siempre tiene todo en True."""
    if usuario.rol == "admin":
        return {clave: True for clave in schemas.CLAVES_PERMISO}
    return _obtener_permisos_cajero(db)



# ============================================================
#  CONFIGURACIÓN DEL NEGOCIO (nombre de la tienda, etc.)
# ============================================================

_CONFIG_DEFAULTS = {
    "nombre_tienda": "Mi Tienda",
}


def _leer_config(db: Session) -> dict:
    """Devuelve la configuración completa, rellenando con defaults."""
    filas = db.query(models.Configuracion).all()
    cfg = dict(_CONFIG_DEFAULTS)
    for f in filas:
        if f.clave in cfg and f.valor is not None and str(f.valor).strip() != "":
            cfg[f.clave] = str(f.valor).strip()
    return cfg


def _nombre_tienda(db: Session) -> str:
    return _leer_config(db)["nombre_tienda"]


@app.get("/api/configuracion", response_model=schemas.ConfiguracionOut)
def obtener_configuracion(db: Session = Depends(get_db)):
    """Pública (sin login): el nombre de la tienda se muestra en el login."""
    return _leer_config(db)


@app.put("/api/configuracion", response_model=schemas.ConfiguracionOut)
def actualizar_configuracion(
    datos: schemas.ConfiguracionActualizar,
    db: Session = Depends(get_db),
    _admin: models.Usuario = Depends(auth.requiere_rol("admin")),
):
    try:
        cambios = datos.dict(exclude_unset=True)
    except Exception:
        cambios = datos.model_dump(exclude_unset=True)

    if "nombre_tienda" in cambios:
        nombre = (cambios["nombre_tienda"] or "").strip()
        if not nombre:
            raise HTTPException(400, "El nombre de la tienda no puede quedar vacío")
        if len(nombre) > 80:
            raise HTTPException(400, "El nombre de la tienda es demasiado largo (máx. 80)")
        fila = db.query(models.Configuracion).get("nombre_tienda")
        if fila:
            fila.valor = nombre
        else:
            db.add(models.Configuracion(clave="nombre_tienda", valor=nombre))
        db.commit()

    return _leer_config(db)


# ============================================================
#  PROVEEDORES  (alta, baja, edición, consulta y sugerencia de pedido)
# ============================================================

DIAS_SEMANA = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]


@app.post("/api/proveedores", response_model=schemas.ProveedorOut)
def crear_proveedor(
    datos: schemas.ProveedorCrear,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("proveedores.agregar")),
):
    if db.query(models.Proveedor).filter(models.Proveedor.nombre == datos.nombre).first():
        raise HTTPException(400, "Ya existe un proveedor con ese nombre")

    nuevo = models.Proveedor(**datos.dict())
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo


@app.get("/api/proveedores", response_model=List[schemas.ProveedorOut])
def listar_proveedores(
    activos: Optional[bool] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_rol("admin", "cajero")),
):
    query = db.query(models.Proveedor)
    if activos is not None:
        query = query.filter(models.Proveedor.activo == activos)
    return query.order_by(models.Proveedor.nombre).all()


@app.put("/api/proveedores/{proveedor_id}", response_model=schemas.ProveedorOut)
def actualizar_proveedor(
    proveedor_id: int,
    datos: schemas.ProveedorActualizar,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("proveedores.editar")),
):
    proveedor = db.query(models.Proveedor).get(proveedor_id)
    if not proveedor:
        raise HTTPException(404, "Proveedor no encontrado")

    cambios = datos.dict(exclude_unset=True)
    if cambios.get("nombre"):
        existente = db.query(models.Proveedor).filter(
            models.Proveedor.nombre == cambios["nombre"],
            models.Proveedor.id != proveedor_id,
        ).first()
        if existente:
            raise HTTPException(400, "Ya existe un proveedor con ese nombre")

    for campo, valor in cambios.items():
        setattr(proveedor, campo, valor)

    db.commit()
    db.refresh(proveedor)
    return proveedor


@app.delete("/api/proveedores/{proveedor_id}")
def dar_de_baja_proveedor(
    proveedor_id: int,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("proveedores.baja")),
):
    """Baja lógica: no se borra, solo se marca como inactivo (sus productos no se tocan)."""
    proveedor = db.query(models.Proveedor).get(proveedor_id)
    if not proveedor:
        raise HTTPException(404, "Proveedor no encontrado")
    proveedor.activo = False
    db.commit()
    return {"ok": True, "mensaje": f"Proveedor '{proveedor.nombre}' dado de baja"}


@app.post("/api/proveedores/{proveedor_id}/reactivar")
def reactivar_proveedor(
    proveedor_id: int,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("proveedores.baja")),
):
    proveedor = db.query(models.Proveedor).get(proveedor_id)
    if not proveedor:
        raise HTTPException(404, "Proveedor no encontrado")
    proveedor.activo = True
    db.commit()
    return {"ok": True}


def _calcular_sugerencia_pedido(db: Session, proveedor: models.Proveedor):
    """
    Lógica de sugerencia de pedido para un proveedor (modelo mínimo/máximo
    dinámico, sin promedios de venta):

    - stock_minimo: lo captura el usuario a mano en cada producto; es el
      nivel en el que hay que reponer.
    - stock_maximo: NO se captura a mano, es dinámico. Cada vez que entra
      mercancía (entrada de inventario, ajuste hacia arriba, alta con stock
      inicial, importación de Excel) y el stock resultante supera el
      máximo que se tenía registrado, ese nuevo nivel se convierte en el
      máximo. Es decir, el máximo siempre es "el nivel de stock más alto
      que ha tenido ese producto".
    - Sugerido a pedir = stock_maximo - stock. Ej.: máximo 20, quedan 5 →
      se sugiere comprar 15 para volver a los 20. Si la próxima vez se
      compran 30, el nuevo máximo pasa a ser 30.
    - Si un producto todavía no tiene un máximo registrado (recién dado de
      alta, nunca tuvo entradas), no hay con qué comparar; mientras tanto
      se sugiere el doble del mínimo como colchón razonable.
    """
    productos = (
        db.query(models.Producto)
        .filter(models.Producto.proveedor_id == proveedor.id, models.Producto.activo == True)
        .order_by(models.Producto.nombre)
        .all()
    )

    items = []
    for p in productos:
        necesita_pedido = p.stock <= p.stock_minimo

        if p.stock_maximo > p.stock_minimo:
            sugerido = p.stock_maximo - p.stock
            motivo = f"Volver al máximo registrado ({p.stock_maximo})"
        else:
            sugerido = (p.stock_minimo * 2) - p.stock
            motivo = "Sin máximo registrado todavía; se sugiere el doble del mínimo"

        # El redondeo depende de cómo se vende el producto: si es por pieza
        # (ej. una Coca-Cola, una bolsa de galletas) no tiene sentido pedir
        # una cantidad fraccionaria, así que se redondea hacia arriba a la
        # pieza completa siguiente. Si es a granel (kg), sí se deja con
        # decimales porque el proveedor puede surtir por kilogramo parcial.
        if p.unidad_venta == "pieza":
            sugerido = math.ceil(sugerido) if sugerido > 0 else 0
        else:
            sugerido = round(sugerido, 3)
        sugerido = max(0, sugerido)

        if necesita_pedido or sugerido > 0:
            items.append(schemas.SugerenciaPedidoItem(
                producto_id=p.id,
                codigo_barras=p.codigo_barras,
                nombre=p.nombre,
                stock=p.stock,
                stock_minimo=p.stock_minimo,
                stock_maximo=p.stock_maximo,
                sugerido=sugerido,
                motivo=motivo,
                costo=p.costo or 0,
                unidad_venta=p.unidad_venta,
            ))

    items.sort(key=lambda i: i.sugerido, reverse=True)
    return items


@app.get("/api/proveedores/{proveedor_id}/sugerencia-pedido", response_model=schemas.SugerenciaPedidoOut)
def sugerencia_pedido(
    proveedor_id: int,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("proveedores.sugerencia")),
):
    proveedor = db.query(models.Proveedor).get(proveedor_id)
    if not proveedor:
        raise HTTPException(404, "Proveedor no encontrado")

    items = _calcular_sugerencia_pedido(db, proveedor)
    return schemas.SugerenciaPedidoOut(
        proveedor=proveedor,
        items=items,
    )


@app.get("/api/proveedores/{proveedor_id}/sugerencia-pedido/excel")
def sugerencia_pedido_excel(
    proveedor_id: int,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("proveedores.sugerencia")),
):
    """Exporta la sugerencia de pedido a Excel, lista para mandarle al proveedor o imprimirla."""
    from openpyxl import Workbook

    proveedor = db.query(models.Proveedor).get(proveedor_id)
    if not proveedor:
        raise HTTPException(404, "Proveedor no encontrado")

    items = _calcular_sugerencia_pedido(db, proveedor)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sugerencia de pedido"
    ws.append([f"Sugerencia de pedido — {proveedor.nombre}"])
    ws.append([f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Código", "Producto", "Stock actual", "Stock mínimo", "Stock máximo", "Sugerido a pedir", "Motivo"])
    from openpyxl.styles import Font, PatternFill
    fila_encabezado = ws[4]
    for celda in fila_encabezado:
        celda.font = Font(color="FFFFFF", bold=True)
        celda.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for it in items:
        ws.append([
            it.codigo_barras, it.nombre, it.stock, it.stock_minimo,
            it.stock_maximo, it.sugerido, it.motivo,
        ])
    _autoajustar_columnas(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre_archivo = f"pedido_{proveedor.nombre.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )


@app.post("/api/proveedores/{proveedor_id}/hacer-pedido", response_model=schemas.PedidoConfirmado)
def hacer_pedido(
    proveedor_id: int,
    pedido: schemas.PedidoCrear,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("proveedores.sugerencia")),
):
    """
    Convierte la sugerencia de pedido en una compra real: por cada producto
    con cantidad > 0 (ya sea la sugerida o la que el dueño haya ajustado a
    mano), registra una entrada de inventario -igual que si se diera de alta
    en la sección de Inventario- lo que sube el stock del producto y, de
    paso, alimenta el reporte de gastos con el costo de esa compra.
    """
    proveedor = db.query(models.Proveedor).get(proveedor_id)
    if not proveedor:
        raise HTTPException(404, "Proveedor no encontrado")

    items_validos = [it for it in pedido.items if it.cantidad and it.cantidad > 0]
    if not items_validos:
        raise HTTPException(400, "Selecciona al menos un producto con una cantidad mayor a 0")

    resultado = []
    total_gasto = 0.0
    total_piezas = 0.0

    for it in items_validos:
        producto = db.query(models.Producto).filter(
            models.Producto.id == it.producto_id,
            models.Producto.proveedor_id == proveedor_id,
        ).first()
        if not producto:
            raise HTTPException(404, f"Uno de los productos ya no pertenece a este proveedor")

        if producto.unidad_venta == "pieza" and it.cantidad != int(it.cantidad):
            raise HTTPException(400, f"'{producto.nombre}' se vende por pieza, la cantidad debe ser un número entero")

        costo_unitario = producto.costo or 0
        subtotal = round(costo_unitario * it.cantidad, 2)

        producto.stock += it.cantidad
        # Mismo criterio que en cualquier otra entrada de inventario: el
        # máximo dinámico sube si este pedido deja el stock por encima del
        # nivel más alto que se tenía registrado.
        if producto.stock > producto.stock_maximo:
            producto.stock_maximo = producto.stock

        db.add(models.MovimientoInventario(
            producto_id=producto.id, tipo="entrada",
            cantidad=it.cantidad,
            costo_unitario=costo_unitario if costo_unitario else None,
            motivo=f"Pedido a proveedor: {proveedor.nombre}",
        ))

        total_gasto += subtotal
        total_piezas += it.cantidad
        resultado.append(schemas.PedidoConfirmadoItem(
            producto_id=producto.id,
            nombre=producto.nombre,
            cantidad=it.cantidad,
            costo_unitario=costo_unitario,
            subtotal=subtotal,
            stock_nuevo=producto.stock,
        ))

    db.commit()

    return schemas.PedidoConfirmado(
        proveedor_id=proveedor_id,
        proveedor_nombre=proveedor.nombre,
        items=resultado,
        total_gasto=round(total_gasto, 2),
        total_productos=len(resultado),
        total_piezas=round(total_piezas, 3),
    )


# ============================================================
#  PRODUCTOS  (alta, baja, edición, consulta)
# ============================================================

UNIDADES_VENTA_VALIDAS = ("pieza", "kg")


def _generar_codigo_interno(db: Session) -> str:
    """
    Genera un código interno único para productos que no tienen código de
    barras real (a granel, o piezas sueltas como huevo/pan). Se usa el
    prefijo 'INT-' para que nunca choque con un código de barras real
    (EAN/UPC son solo números) y para poder identificarlos fácilmente.
    El cajero nunca ve ni escribe este código: el producto se vende desde
    los botones de "Venta rápida".
    """
    while True:
        candidato = "INT-" + secrets.token_hex(4).upper()
        if not db.query(models.Producto).filter(models.Producto.codigo_barras == candidato).first():
            return candidato


@app.post("/api/productos", response_model=schemas.ProductoOut)
def crear_producto(
    producto: schemas.ProductoCrear,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("productos.agregar")),
):
    if producto.unidad_venta not in UNIDADES_VENTA_VALIDAS:
        raise HTTPException(400, "unidad_venta debe ser 'pieza' o 'kg'")

    datos = producto.dict()

    if datos.get("proveedor_id") is not None:
        if not db.query(models.Proveedor).get(datos["proveedor_id"]):
            raise HTTPException(400, "El proveedor indicado no existe")

    if not datos["requiere_codigo"]:
        # Producto sin código de barras real (granel o pieza suelta):
        # ignoramos cualquier código que hayan mandado y generamos uno interno.
        datos["codigo_barras"] = _generar_codigo_interno(db)
    else:
        if not datos.get("codigo_barras"):
            raise HTTPException(400, "El código de barras es obligatorio si el producto sí tiene código")
        existente = db.query(models.Producto).filter(
            models.Producto.codigo_barras == datos["codigo_barras"]
        ).first()
        if existente:
            raise HTTPException(400, "Ya existe un producto con ese código de barras")

    nuevo = models.Producto(**datos)
    # El máximo arranca igual al stock inicial (o al mínimo si no se cargó
    # stock todavía); de ahí en adelante sube solo con cada entrada que lo
    # supere.
    nuevo.stock_maximo = max(nuevo.stock, nuevo.stock_minimo)
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)

    if nuevo.stock > 0:
        mov = models.MovimientoInventario(
            producto_id=nuevo.id, tipo="entrada",
            cantidad=nuevo.stock, motivo="Alta inicial de producto",
            # Se guarda el costo capturado en el alta para que este stock
            # inicial sí cuente en el reporte de "Gastado en compras", igual
            # que cualquier otra entrada.
            costo_unitario=nuevo.costo if nuevo.costo else None,
        )
        db.add(mov)
        db.commit()

    return nuevo


@app.get("/api/productos", response_model=List[schemas.ProductoOut])
def listar_productos(
    activos: Optional[bool] = None,
    buscar: Optional[str] = None,
    requiere_codigo: Optional[bool] = None,
    proveedor_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_rol("admin", "cajero")),
):
    query = db.query(models.Producto)
    if activos is not None:
        query = query.filter(models.Producto.activo == activos)
    if proveedor_id is not None:
        query = query.filter(models.Producto.proveedor_id == proveedor_id)
    if requiere_codigo is not None:
        # Usado por la pantalla de "Venta rápida" para traer solo los
        # productos sin código de barras (a granel / piezas sueltas).
        query = query.filter(models.Producto.requiere_codigo == requiere_codigo)
    if buscar:
        like = f"%{buscar}%"
        query = query.filter(
            (models.Producto.nombre.ilike(like)) |
            (models.Producto.codigo_barras.ilike(like))
        )
    return query.order_by(models.Producto.nombre).all()



@app.get("/api/productos/venta-rapida", response_model=List[schemas.ProductoOut])
def productos_venta_rapida(
    limite: int = 9,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_rol("admin", "cajero")),
):
    """
    Atajos en Vender: los productos más vendidos del catálogo completo
    (con código de barras y sin él).

    Devuelve hasta `limite` productos activos (por defecto 9), priorizando los
    más vendidos (unidades en ventas no canceladas). Si hay menos de `limite`
    con historial, se completan con el resto (por nombre) para que una tienda
    nueva no vea la cuadrícula vacía.
    """
    limite = max(1, min(int(limite or 9), 24))

    base = db.query(models.Producto).filter(
        models.Producto.activo == True,
    )
    if base.count() == 0:
        return []

    # Unidades vendidas por producto (solo ventas vigentes), de todo el catálogo.
    ranking = (
        db.query(
            models.VentaDetalle.producto_id,
            func.sum(models.VentaDetalle.cantidad).label("unidades"),
        )
        .join(models.Venta, models.Venta.id == models.VentaDetalle.venta_id)
        .join(models.Producto, models.Producto.id == models.VentaDetalle.producto_id)
        .filter(
            models.Venta.cancelada == False,
            models.Producto.activo == True,
        )
        .group_by(models.VentaDetalle.producto_id)
        .order_by(func.sum(models.VentaDetalle.cantidad).desc())
        .all()
    )

    ids_orden = []
    vistos = set()
    for fila in ranking:
        if fila.producto_id in vistos:
            continue
        ids_orden.append(fila.producto_id)
        vistos.add(fila.producto_id)
        if len(ids_orden) >= limite:
            break

    # Completar con productos sin (o con poca) venta, por nombre.
    if len(ids_orden) < limite:
        q_resto = base.order_by(models.Producto.nombre.asc())
        if vistos:
            q_resto = q_resto.filter(~models.Producto.id.in_(list(vistos)))
        resto = q_resto.limit(limite - len(ids_orden)).all()
        for p in resto:
            ids_orden.append(p.id)
            vistos.add(p.id)

    productos = db.query(models.Producto).filter(models.Producto.id.in_(ids_orden)).all()
    por_id = {p.id: p for p in productos}
    return [por_id[i] for i in ids_orden if i in por_id]


@app.get("/api/productos/codigo/{codigo_barras}", response_model=schemas.ProductoOut)
def buscar_por_codigo(
    codigo_barras: str,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_rol("admin", "cajero")),
):
    """Usado por el escáner del celular para encontrar el producto al vuelo."""
    producto = db.query(models.Producto).filter(
        models.Producto.codigo_barras == codigo_barras
    ).first()
    if not producto:
        raise HTTPException(404, "Producto no encontrado")
    return producto


# Bases de datos públicas de códigos de barras, en orden de prioridad.
# Se consultan solo cuando el producto NO existe en la base local
# (el escaneo local/offline sigue siendo el método principal y no depende de esto).
_FUENTES_CODIGO_BARRAS = [
    ("openfoodfacts", "https://world.openfoodfacts.org/api/v2/product/{codigo}.json"),
    ("openbeautyfacts", "https://world.openbeautyfacts.org/api/v2/product/{codigo}.json"),
    ("openproductsfacts", "https://world.openproductsfacts.org/api/v2/product/{codigo}.json"),
]


@app.get("/api/productos/buscar-web/{codigo_barras}")
def buscar_en_web(
    codigo_barras: str,
    _usuario: models.Usuario = Depends(auth.requiere_rol("admin", "cajero")),
):
    """
    Busca el código de barras en bases de datos públicas de internet
    (Open Food/Beauty/Products Facts) para sugerir nombre, marca y
    categoría al dar de alta un producto nuevo.

    Si no hay internet, la fuente no responde a tiempo, o el código
    no está registrado en ninguna, devuelve encontrado=False en vez
    de fallar, para que el frontend siga el flujo de alta manual
    (segundo método, que no depende de la web).
    """
    with httpx.Client(timeout=4.0) as cliente:
        for _nombre_fuente, url_plantilla in _FUENTES_CODIGO_BARRAS:
            try:
                resp = cliente.get(url_plantilla.format(codigo=codigo_barras))
                if resp.status_code != 200:
                    continue
                data = resp.json()
                if data.get("status") != 1:
                    continue
                producto = data.get("product", {})
                nombre = (
                    producto.get("product_name_es")
                    or producto.get("product_name")
                    or producto.get("generic_name")
                )
                if not nombre:
                    continue
                marca = producto.get("brands", "")
                nombre_completo = f"{nombre} ({marca})" if marca else nombre
                return {
                    "encontrado": True,
                    "codigo_barras": codigo_barras,
                    "nombre": nombre_completo,
                    "categoria": (producto.get("categories") or "General").split(",")[0].strip(),
                    "imagen": producto.get("image_front_small_url") or producto.get("image_url"),
                }
            except (httpx.TimeoutException, httpx.RequestError):
                # Sin internet o la fuente falló: probamos la siguiente,
                # y si todas fallan devolvemos encontrado=False abajo.
                continue

    return {"encontrado": False, "codigo_barras": codigo_barras}


@app.put("/api/productos/{producto_id}", response_model=schemas.ProductoOut)
def actualizar_producto(
    producto_id: int,
    datos: schemas.ProductoActualizar,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("productos.editar")),
):
    producto = db.query(models.Producto).get(producto_id)
    if not producto:
        raise HTTPException(404, "Producto no encontrado")

    cambios = datos.dict(exclude_unset=True)
    if cambios.get("proveedor_id") is not None:
        if not db.query(models.Proveedor).get(cambios["proveedor_id"]):
            raise HTTPException(400, "El proveedor indicado no existe")

    for campo, valor in cambios.items():
        setattr(producto, campo, valor)

    db.commit()
    db.refresh(producto)
    return producto


@app.delete("/api/productos/{producto_id}")
def dar_de_baja(
    producto_id: int,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("productos.baja")),
):
    """Baja lógica: no se borra de la BD, solo se marca como inactivo."""
    producto = db.query(models.Producto).get(producto_id)
    if not producto:
        raise HTTPException(404, "Producto no encontrado")
    producto.activo = False
    db.commit()
    return {"ok": True, "mensaje": f"Producto '{producto.nombre}' dado de baja"}


@app.post("/api/productos/{producto_id}/reactivar")
def reactivar_producto(
    producto_id: int,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("productos.baja")),
):
    producto = db.query(models.Producto).get(producto_id)
    if not producto:
        raise HTTPException(404, "Producto no encontrado")
    producto.activo = True
    db.commit()
    return {"ok": True}


# ------------------------------------------------------------
#  Importar / exportar catálogo de productos por Excel
# ------------------------------------------------------------

_ENCABEZADOS_PRODUCTOS = [
    "Código de barras", "¿Tiene código de barras?", "Nombre", "Descripción", "Categoría",
    "Precio de venta", "Costo", "Stock", "Stock mínimo", "Unidad de venta", "Proveedor",
]

# Alias aceptados para cada columna al leer un Excel importado
# (para tolerar mayúsculas/minúsculas, acentos, o nombres ligeramente distintos).
_ALIAS_COLUMNAS_PRODUCTOS = {
    "codigo_barras": ["código de barras", "codigo de barras", "código", "codigo", "barcode"],
    # Columna explícita para decidir si el producto se vende con código de
    # barras real (se escanea) o a granel/pieza suelta (aparece en Venta
    # rápida). Valores esperados: "Sí" / "No". Si se deja en blanco o la
    # columna no viene en el Excel, se deduce del código de barras: si la
    # fila no trae código, se asume "No" (a granel/sin código).
    "tiene_codigo": ["¿tiene código de barras?", "tiene código de barras", "tiene codigo de barras",
                     "requiere código de barras", "requiere codigo de barras", "con código"],
    "nombre": ["nombre", "producto"],
    "descripcion": ["descripción", "descripcion"],
    "categoria": ["categoría", "categoria"],
    "precio_venta": ["precio de venta", "precio venta", "precio"],
    "costo": ["costo"],
    "stock": ["stock", "existencia", "existencias"],
    "stock_minimo": ["stock mínimo", "stock minimo", "stock min"],
    # Valores esperados: "Pieza" o "Kg" (si se deja en blanco, se asume "Pieza").
    "unidad_venta": ["unidad de venta", "unidad", "u. venta"],
    # Nombre del proveedor. Si no existe todavía, se da de alta
    # automáticamente al importar (no bloquea la importación del catálogo).
    "proveedor": ["proveedor", "proveedor sugerido"],
}


def _estilizar_encabezado(hoja):
    from openpyxl.styles import Font, PatternFill
    fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for celda in hoja[1]:
        celda.font = font
        celda.fill = fill


def _autoajustar_columnas(hoja):
    for columna in hoja.columns:
        ancho = max((len(str(c.value)) for c in columna if c.value is not None), default=10)
        hoja.column_dimensions[columna[0].column_letter].width = min(ancho + 4, 40)


@app.get("/api/productos/plantilla-excel")
def plantilla_excel_productos(
    _usuario: models.Usuario = Depends(auth.requiere_permiso("productos.importar")),
):
    """Descarga una plantilla en blanco (con un ejemplo) para importar productos."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(_ENCABEZADOS_PRODUCTOS)
    ws.append(["7501234567890", "Sí", "Producto de ejemplo", "Descripción opcional", "General", 25.50, 15.00, 10, 2, "Pieza", "Coca-Cola FEMSA"])
    ws.append(["", "No", "Producto a granel de ejemplo", "Se vende suelto, sin escanear", "General", 18.00, 10.00, 5, 1, "Kg", ""])
    _estilizar_encabezado(ws)
    _autoajustar_columnas(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos.xlsx"},
    )


@app.get("/api/productos/exportar/excel")
def exportar_productos_excel(
    activos: Optional[bool] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("productos.importar", "productos.ver")),
):
    """Exporta el catálogo completo de productos a un archivo Excel."""
    from openpyxl import Workbook

    query = db.query(models.Producto)
    if activos is not None:
        query = query.filter(models.Producto.activo == activos)
    productos = query.order_by(models.Producto.nombre).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(_ENCABEZADOS_PRODUCTOS + ["Activo"])
    for p in productos:
        ws.append([
            p.codigo_barras,
            "Sí" if p.requiere_codigo else "No",
            p.nombre, p.descripcion or "", p.categoria or "",
            p.precio_venta, p.costo, p.stock, p.stock_minimo,
            "Kg" if p.unidad_venta == "kg" else "Pieza",
            p.proveedor.nombre if p.proveedor else "",
            "Sí" if p.activo else "No",
        ])
    _estilizar_encabezado(ws)
    _autoajustar_columnas(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=productos.xlsx"},
    )


@app.post("/api/productos/importar/excel")
def importar_productos_excel(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("productos.importar")),
):
    """
    Importa productos desde un Excel (.xlsx). Si el código de barras ya
    existe se actualiza el producto; si no existe, se crea uno nuevo.
    Se identifican las columnas por su nombre de encabezado (fila 1),
    así que el orden de las columnas no importa.
    """
    from openpyxl import load_workbook

    if not archivo.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "El archivo debe ser un Excel (.xlsx)")

    try:
        contenido = archivo.file.read()
        wb = load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "No se pudo leer el archivo. Verifica que sea un Excel válido")

    encabezados = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
    indices = {}
    for campo, alias in _ALIAS_COLUMNAS_PRODUCTOS.items():
        for i, encabezado in enumerate(encabezados):
            if encabezado in alias:
                indices[campo] = i
                break

    if not all(c in indices for c in ("codigo_barras", "nombre", "precio_venta")):
        raise HTTPException(
            400,
            "El Excel debe tener al menos las columnas: 'Código de barras', 'Nombre' y 'Precio de venta'. "
            "Descarga la plantilla para ver el formato correcto.",
        )

    creados, actualizados, errores = 0, 0, []

    for num_fila, fila in enumerate(ws.iter_rows(min_row=2), start=2):
        def valor(campo, por_defecto=None):
            idx = indices.get(campo)
            if idx is None or idx >= len(fila):
                return por_defecto
            v = fila[idx].value
            return v if v is not None else por_defecto

        codigo = str(valor("codigo_barras", "") or "").strip()
        nombre = str(valor("nombre", "") or "").strip()

        if not codigo and not nombre:
            continue  # fila totalmente vacía, se ignora

        if not nombre:
            errores.append(f"Fila {num_fila}: falta el nombre")
            continue

        tiene_codigo_texto = valor("tiene_codigo", None)
        if tiene_codigo_texto is not None:
            tiene_codigo_texto = str(tiene_codigo_texto).strip().lower()

        if tiene_codigo_texto in ("sí", "si", "s", "true", "1", "yes"):
            if not codigo:
                errores.append(f"Fila {num_fila}: marcaste 'Sí' en '¿Tiene código de barras?' pero no pusiste el código")
                continue
            sin_codigo_real = False
        elif tiene_codigo_texto in ("no", "n", "false", "0"):
            sin_codigo_real = True
        else:
            # Columna no incluida en el Excel o celda vacía: se deduce del código.
            sin_codigo_real = not codigo

        if sin_codigo_real and not codigo:
            # Producto sin código de barras real (a granel, pieza suelta, etc.):
            # se le asigna un código interno, igual que al darlo de alta manualmente.
            codigo = _generar_codigo_interno(db)

        try:
            precio = float(valor("precio_venta", 0) or 0)
            costo = float(valor("costo", 0) or 0)
            stock_minimo = float(valor("stock_minimo", 0) or 0)
        except (TypeError, ValueError):
            errores.append(f"Fila {num_fila}: precio, costo o stock mínimo inválido")
            continue

        unidad_texto = valor("unidad_venta", None)
        unidad_venta = None
        if unidad_texto is not None:
            unidad_texto = str(unidad_texto).strip().lower()
            unidad_venta = "kg" if unidad_texto in ("kg", "kilogramo", "kilo", "granel") else "pieza"

        stock_excel = valor("stock", None)
        try:
            stock_excel = float(stock_excel) if stock_excel is not None else None
        except (TypeError, ValueError):
            errores.append(f"Fila {num_fila}: el stock debe ser un número")
            continue

        datos_comunes = {
            "nombre": nombre,
            "descripcion": str(valor("descripcion", "") or ""),
            "categoria": str(valor("categoria", "General") or "General"),
            "precio_venta": precio,
            "costo": costo,
            "stock_minimo": stock_minimo,
            # Si la fila no traía código de barras, el producto es "sin código
            # real" (a granel/pieza suelta): debe aparecer en Venta rápida en
            # vez de esperarse a que lo escaneen con su código interno.
            "requiere_codigo": not sin_codigo_real,
        }
        # Solo se toca unidad_venta si la columna vino en el Excel; si no,
        # se respeta la que ya tenía el producto (o el default 'pieza' al crear uno nuevo).
        if unidad_venta is not None:
            datos_comunes["unidad_venta"] = unidad_venta

        # Igual que unidad_venta: solo se toca proveedor_id si la columna
        # "Proveedor" vino en el Excel. Si la celda viene vacía se quita el
        # proveedor del producto; si trae un nombre que no existe todavía,
        # se da de alta el proveedor automáticamente (frecuencia semanal por
        # defecto, editable después desde el módulo de Proveedores).
        if "proveedor" in indices:
            nombre_proveedor = str(valor("proveedor", "") or "").strip()
            if nombre_proveedor:
                proveedor = db.query(models.Proveedor).filter(
                    func.lower(models.Proveedor.nombre) == nombre_proveedor.lower()
                ).first()
                if not proveedor:
                    proveedor = models.Proveedor(nombre=nombre_proveedor)
                    db.add(proveedor)
                    db.flush()
                datos_comunes["proveedor_id"] = proveedor.id
            else:
                datos_comunes["proveedor_id"] = None

        producto = db.query(models.Producto).filter(models.Producto.codigo_barras == codigo).first()

        if producto:
            for campo, val in datos_comunes.items():
                setattr(producto, campo, val)
            if stock_excel is not None and stock_excel != producto.stock:
                diferencia = stock_excel - producto.stock
                producto.stock = stock_excel
                db.add(models.MovimientoInventario(
                    producto_id=producto.id, tipo="ajuste",
                    cantidad=diferencia, motivo="Ajuste por importación de Excel",
                ))
            if producto.stock > producto.stock_maximo:
                producto.stock_maximo = producto.stock
            actualizados += 1
        else:
            nuevo = models.Producto(codigo_barras=codigo, stock=stock_excel or 0, **datos_comunes)
            nuevo.stock_maximo = max(nuevo.stock, nuevo.stock_minimo)
            db.add(nuevo)
            db.flush()
            if nuevo.stock > 0:
                db.add(models.MovimientoInventario(
                    producto_id=nuevo.id, tipo="entrada",
                    cantidad=nuevo.stock, motivo="Alta por importación de Excel",
                    costo_unitario=nuevo.costo if nuevo.costo else None,
                ))
            creados += 1

    db.commit()
    return {"creados": creados, "actualizados": actualizados, "errores": errores}


# ============================================================
#  INVENTARIO
# ============================================================

@app.post("/api/inventario/movimiento", response_model=schemas.MovimientoOut)
def registrar_movimiento(
    mov: schemas.MovimientoCrear,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("inventario.movimiento")),
):
    producto = db.query(models.Producto).filter(
        models.Producto.codigo_barras == mov.codigo_barras
    ).first()
    if not producto:
        raise HTTPException(404, "Producto no encontrado")

    if producto.unidad_venta == "pieza" and mov.cantidad != int(mov.cantidad):
        raise HTTPException(400, f"'{producto.nombre}' se vende por pieza, la cantidad debe ser un número entero")

    if mov.tipo in ("entrada", "salida"):
        if mov.cantidad <= 0:
            raise HTTPException(400, "La cantidad debe ser mayor a 0")
    elif mov.tipo == "ajuste":
        # El ajuste FIJA el stock al valor indicado: no puede ser negativo.
        if mov.cantidad < 0:
            raise HTTPException(400, "El stock de un ajuste no puede ser negativo")
    else:
        raise HTTPException(400, "Tipo de movimiento inválido (entrada/salida/ajuste)")

    costo_unitario_mov = None
    if mov.tipo == "entrada":
        producto.stock += mov.cantidad
        # Si no llega el costo de esta compra (ej. un cliente que no manda
        # el campo), se usa el costo ya registrado del producto en vez de
        # dejarlo vacío, para que el movimiento sí cuente en el reporte de
        # gastos. Si sí llega, se actualiza el costo "actual" del producto
        # con el de esta compra, ya que puede variar de una a otra.
        costo_unitario_mov = mov.costo_unitario
        if costo_unitario_mov is not None:
            producto.costo = costo_unitario_mov
        elif producto.costo:
            costo_unitario_mov = producto.costo
    elif mov.tipo == "salida":
        if producto.stock < mov.cantidad:
            raise HTTPException(400, "Stock insuficiente para esa salida")
        producto.stock -= mov.cantidad
        # La salida (merma/pérdida) se valoriza con el costo actual del
        # producto, para poder verla como dinero perdido en el reporte de
        # mermas. A diferencia de la entrada, aquí NUNCA se actualiza el
        # costo del producto, porque no es una compra.
        if producto.costo:
            costo_unitario_mov = producto.costo
    elif mov.tipo == "ajuste":
        producto.stock = mov.cantidad  # ajuste fija el stock al valor indicado

    # El máximo es dinámico: si el stock resultante supera el máximo que
    # se tenía registrado, ese nuevo nivel se convierte en el máximo. Así
    # se captura solo, sin que el usuario tenga que definirlo a mano.
    if producto.stock > producto.stock_maximo:
        producto.stock_maximo = producto.stock

    nuevo_mov = models.MovimientoInventario(
        producto_id=producto.id, tipo=mov.tipo,
        cantidad=mov.cantidad,
        costo_unitario=costo_unitario_mov,
        motivo=mov.motivo
    )
    db.add(nuevo_mov)
    db.commit()
    db.refresh(nuevo_mov)
    return nuevo_mov


@app.get("/api/inventario/movimientos", response_model=List[schemas.MovimientoOut])
def listar_movimientos(
    producto_id: Optional[int] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("inventario.ver")),
):
    query = db.query(models.MovimientoInventario)
    if producto_id:
        query = query.filter(models.MovimientoInventario.producto_id == producto_id)
    if fecha_inicio:
        query = query.filter(models.MovimientoInventario.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.MovimientoInventario.fecha < fecha_fin + timedelta(days=1))
    return query.order_by(models.MovimientoInventario.fecha.desc()).limit(500).all()


@app.get("/api/inventario/bajo-stock", response_model=List[schemas.ProductoOut])
def productos_bajo_stock(
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("inventario.ver")),
):
    return db.query(models.Producto).filter(
        models.Producto.activo == True,
        models.Producto.stock <= models.Producto.stock_minimo
    ).all()


# ============================================================
#  NOTIFICACIONES EN TIEMPO REAL (dashboard / campana)
# ============================================================
#
# No hay una tabla de "notificaciones": se calculan al vuelo, cada vez que
# se piden, a partir de la configuración que ya existe (día de visita de
# cada proveedor, stock mínimo de cada producto). Así siempre reflejan el
# estado actual del negocio. Lo único que se guarda es qué avisos ya
# descartó el usuario (tabla notificaciones_descartadas), para no volver
# a mostrarlos.

def _puede_ver_modulo(db: Session, usuario: models.Usuario, clave: str) -> bool:
    """Como auth.requiere_permiso, pero sin lanzar excepción (regresa bool)."""
    if usuario.rol == "admin":
        return True
    permisos = _obtener_permisos_cajero(db)
    return bool(permisos.get(clave))


def _generar_notificaciones(db: Session, usuario: models.Usuario) -> List[schemas.NotificacionOut]:
    notificaciones: List[schemas.NotificacionOut] = []
    ahora = datetime.now()
    hoy = ahora.date()

    descartadas = {c for (c,) in db.query(models.NotificacionDescartada.clave).all()}

    # --- Proveedores que pasan hoy ---------------------------------
    if _puede_ver_modulo(db, usuario, "proveedores.ver"):
        dia_semana = hoy.weekday()  # 0=lunes ... 6=domingo (igual que dia_visita)
        proveedores_hoy = (
            db.query(models.Proveedor)
            .filter(models.Proveedor.activo == True, models.Proveedor.dia_visita == dia_semana)
            .order_by(models.Proveedor.nombre)
            .all()
        )
        for p in proveedores_hoy:
            clave = f"proveedor_visita:{p.id}:{hoy.isoformat()}"
            if clave in descartadas:
                continue
            mensaje = f"Hoy es el día en que suele pasar {p.nombre} a levantar el pedido."
            if p.contacto:
                mensaje += f" Contacto: {p.contacto}"
            if p.telefono:
                mensaje += f" ({p.telefono})"
            notificaciones.append(schemas.NotificacionOut(
                clave=clave,
                tipo="proveedor_visita",
                icono="🚚",
                titulo=f"Hoy pasa el proveedor {p.nombre}",
                mensaje=mensaje,
                tab_destino="lista-proveedores",
                texto_boton="Ver lista de proveedores",
                prioridad="info",
                fecha=ahora,
            ))

        # Limpieza ligera: los avisos de visita de proveedor de hace más de
        # 3 días ya no le sirven a nadie (nunca se van a volver a mostrar
        # porque incluyen la fecha en la clave), se borran para no acumular
        # basura en la tabla.
        limite = hoy - timedelta(days=3)
        db.query(models.NotificacionDescartada).filter(
            models.NotificacionDescartada.clave.like("proveedor_visita:%"),
            models.NotificacionDescartada.fecha < datetime.combine(limite, datetime.min.time()),
        ).delete(synchronize_session=False)

    # --- Productos con stock bajo -----------------------------------
    if _puede_ver_modulo(db, usuario, "inventario.ver"):
        productos_bajo = (
            db.query(models.Producto)
            .filter(models.Producto.activo == True, models.Producto.stock <= models.Producto.stock_minimo)
            .order_by(models.Producto.nombre)
            .all()
        )
        for p in productos_bajo:
            clave = f"stock_bajo:{p.id}"
            if clave in descartadas:
                continue
            agotado = p.stock <= 0
            unidad = "kg" if p.unidad_venta == "kg" else "pza"
            notificaciones.append(schemas.NotificacionOut(
                clave=clave,
                tipo="stock_bajo",
                icono="🚨" if agotado else "⚠️",
                titulo=f"{'Sin existencias' if agotado else 'Stock bajo'}: {p.nombre}",
                mensaje=(
                    f"Quedan {p.stock:g} {unidad} (mínimo configurado: {p.stock_minimo:g} {unidad})."
                    if not agotado else
                    f"Se acabó el stock de {p.nombre} (mínimo configurado: {p.stock_minimo:g} {unidad})."
                ),
                tab_destino="alertas-stock",
                texto_boton="Ver alertas de stock",
                prioridad="alta" if agotado else "media",
                fecha=ahora,
            ))

        # Auto-limpieza: si un producto ya se recuperó (stock por arriba del
        # mínimo) y su aviso seguía descartado, se borra el descarte para
        # que, si el stock vuelve a bajar más adelante, la alerta "renazca"
        # en vez de quedar escondida para siempre.
        ids_bajo = {p.id for p in productos_bajo}
        for d in db.query(models.NotificacionDescartada).filter(
            models.NotificacionDescartada.clave.like("stock_bajo:%")
        ).all():
            try:
                pid = int(d.clave.split(":", 1)[1])
            except (IndexError, ValueError):
                continue
            if pid not in ids_bajo:
                db.delete(d)

    db.commit()

    orden_prioridad = {"alta": 0, "media": 1, "info": 2}
    notificaciones.sort(key=lambda n: (orden_prioridad.get(n.prioridad, 9), n.titulo))
    return notificaciones


@app.get("/api/notificaciones", response_model=List[schemas.NotificacionOut])
def obtener_notificaciones(
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(auth.obtener_usuario_actual),
):
    return _generar_notificaciones(db, usuario)


@app.post("/api/notificaciones/descartar")
def descartar_notificacion(
    datos: schemas.NotificacionDescartar,
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(auth.obtener_usuario_actual),
):
    if not datos.clave or not datos.clave.strip():
        raise HTTPException(400, "Falta la clave de la notificación")
    ya_existe = (
        db.query(models.NotificacionDescartada)
        .filter(models.NotificacionDescartada.clave == datos.clave)
        .first()
    )
    if not ya_existe:
        db.add(models.NotificacionDescartada(clave=datos.clave, usuario_id=usuario.id))
        db.commit()
    return {"ok": True}


# ============================================================
#  VENTAS
# ============================================================

@app.post("/api/ventas", response_model=schemas.VentaOut)
def registrar_venta(
    venta: schemas.VentaCrear,
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(auth.requiere_rol("admin", "cajero")),
):
    if not venta.items:
        raise HTTPException(400, "La venta no tiene productos")

    nueva_venta = models.Venta(total=0.0, metodo_pago=venta.metodo_pago, usuario_id=usuario.id)
    db.add(nueva_venta)
    db.flush()  # obtiene el id sin cerrar la transacción

    total = 0.0
    for item in venta.items:
        producto = db.query(models.Producto).filter(
            models.Producto.codigo_barras == item.codigo_barras
        ).first()
        if not producto:
            db.rollback()
            raise HTTPException(404, f"Producto con código {item.codigo_barras} no encontrado")
        if not producto.activo:
            db.rollback()
            raise HTTPException(400, f"El producto '{producto.nombre}' está dado de baja")
        if producto.unidad_venta == "pieza" and item.cantidad != int(item.cantidad):
            db.rollback()
            raise HTTPException(400, f"'{producto.nombre}' se vende por pieza, no se puede vender una cantidad fraccionaria")
        if producto.stock < item.cantidad:
            db.rollback()
            raise HTTPException(400, f"Stock insuficiente de '{producto.nombre}' (disponible: {producto.stock})")

        subtotal = producto.precio_venta * item.cantidad
        detalle = models.VentaDetalle(
            venta_id=nueva_venta.id,
            producto_id=producto.id,
            cantidad=item.cantidad,
            precio_unitario=producto.precio_venta,
            subtotal=subtotal,
        )
        db.add(detalle)

        producto.stock -= item.cantidad
        db.add(models.MovimientoInventario(
            producto_id=producto.id, tipo="venta",
            cantidad=item.cantidad, motivo=f"Venta #{nueva_venta.id}"
        ))

        total += subtotal

    nueva_venta.total = total
    db.commit()
    db.refresh(nueva_venta)

    # Adjunta el nombre del producto en cada detalle para la respuesta
    detalles_out = []
    for d in nueva_venta.detalles:
        detalles_out.append(schemas.VentaDetalleOut(
            producto_id=d.producto_id,
            nombre_producto=d.producto.nombre,
            cantidad=d.cantidad,
            precio_unitario=d.precio_unitario,
            subtotal=d.subtotal,
        ))

    return schemas.VentaOut(
        id=nueva_venta.id, fecha=nueva_venta.fecha, total=nueva_venta.total,
        metodo_pago=nueva_venta.metodo_pago, cancelada=nueva_venta.cancelada,
        vendido_por=usuario.nombre_completo or usuario.username,
        detalles=detalles_out
    )


@app.get("/api/ventas", response_model=List[schemas.VentaOut])
def listar_ventas(
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("reportes.ver")),
):
    query = db.query(models.Venta)
    if fecha_inicio:
        query = query.filter(models.Venta.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.Venta.fecha < fecha_fin + timedelta(days=1))
    ventas = query.order_by(models.Venta.fecha.desc()).all()

    resultado = []
    for v in ventas:
        detalles_out = [
            schemas.VentaDetalleOut(
                producto_id=d.producto_id,
                nombre_producto=d.producto.nombre if d.producto else "—",
                cantidad=d.cantidad,
                precio_unitario=d.precio_unitario,
                subtotal=d.subtotal,
            ) for d in v.detalles
        ]
        resultado.append(schemas.VentaOut(
            id=v.id, fecha=v.fecha, total=v.total,
            metodo_pago=v.metodo_pago, cancelada=v.cancelada,
            vendido_por=(v.usuario.nombre_completo or v.usuario.username) if v.usuario else None,
            detalles=detalles_out
        ))
    return resultado


@app.post("/api/ventas/{venta_id}/cancelar")
def cancelar_venta(
    venta_id: int,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("reportes.cancelar")),
):
    """Cancela una venta y regresa el stock al inventario."""
    venta = db.query(models.Venta).get(venta_id)
    if not venta:
        raise HTTPException(404, "Venta no encontrada")
    if venta.cancelada:
        raise HTTPException(400, "La venta ya estaba cancelada")

    for d in venta.detalles:
        d.producto.stock += d.cantidad
        db.add(models.MovimientoInventario(
            producto_id=d.producto_id, tipo="entrada",
            cantidad=d.cantidad, motivo=f"Cancelación de venta #{venta.id}"
        ))

    venta.cancelada = True
    db.commit()
    return {"ok": True, "mensaje": "Venta cancelada y stock restituido"}


# ============================================================
#  OTROS GASTOS (renta, luz, sueldos, etc.)
# ============================================================

@app.post("/api/otros-gastos", response_model=schemas.OtroGastoOut)
def crear_otro_gasto(
    datos: schemas.OtroGastoCrear,
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(auth.requiere_permiso("gastos.agregar")),
):
    nuevo = models.OtroGasto(
        concepto=datos.concepto,
        categoria=datos.categoria or "otro",
        monto=datos.monto,
        fecha=datetime.combine(datos.fecha, datetime.min.time()) if datos.fecha else datetime.now(),
        notas=datos.notas or "",
        usuario_id=usuario.id,
    )
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo


@app.get("/api/otros-gastos", response_model=List[schemas.OtroGastoOut])
def listar_otros_gastos(
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    categoria: Optional[str] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("gastos.ver", "reportes.ver")),
):
    query = db.query(models.OtroGasto)
    if fecha_inicio:
        query = query.filter(models.OtroGasto.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.OtroGasto.fecha < fecha_fin + timedelta(days=1))
    if categoria:
        query = query.filter(models.OtroGasto.categoria == categoria)
    return query.order_by(models.OtroGasto.fecha.desc()).all()


@app.delete("/api/otros-gastos/{gasto_id}")
def eliminar_otro_gasto(
    gasto_id: int,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("gastos.eliminar")),
):
    gasto = db.query(models.OtroGasto).get(gasto_id)
    if not gasto:
        raise HTTPException(404, "Gasto no encontrado")
    db.delete(gasto)
    db.commit()
    return {"ok": True, "mensaje": "Gasto eliminado"}


# ============================================================
#  CORTE DE CAJA
# ============================================================

def _totales_ventas_dia(db: Session, dia: date) -> dict:
    """Suma ventas no canceladas de un día, separadas por método de pago."""
    inicio = datetime.combine(dia, datetime.min.time())
    fin = inicio + timedelta(days=1)
    ventas = (
        db.query(models.Venta)
        .filter(
            models.Venta.cancelada == False,
            models.Venta.fecha >= inicio,
            models.Venta.fecha < fin,
        )
        .all()
    )
    total_efectivo = total_tarjeta = total_transferencia = 0.0
    for v in ventas:
        monto = v.total or 0
        metodo = (v.metodo_pago or "efectivo").lower()
        if metodo == "tarjeta":
            total_tarjeta += monto
        elif metodo == "transferencia":
            total_transferencia += monto
        else:
            total_efectivo += monto
    total_ventas = total_efectivo + total_tarjeta + total_transferencia
    return {
        "num_ventas": len(ventas),
        "total_ventas": round(total_ventas, 2),
        "total_efectivo": round(total_efectivo, 2),
        "total_tarjeta": round(total_tarjeta, 2),
        "total_transferencia": round(total_transferencia, 2),
    }


@app.get("/api/caja/precorte", response_model=schemas.PrecorteOut)
def precorte_caja(
    fecha: Optional[date] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("caja.cortar", "caja.ver")),
):
    """
    Calcula los totales del día según el sistema, para que el cajero
    sepa cuánto efectivo debería haber antes de contar billetes.
    """
    dia = fecha or date.today()
    totales = _totales_ventas_dia(db, dia)
    inicio = datetime.combine(dia, datetime.min.time())
    fin = inicio + timedelta(days=1)
    ya_hay = (
        db.query(models.CorteCaja)
        .filter(models.CorteCaja.fecha_corte >= inicio, models.CorteCaja.fecha_corte < fin)
        .count()
        > 0
    )
    return schemas.PrecorteOut(fecha=dia, ya_hay_corte=ya_hay, **totales)


@app.post("/api/caja/cortes", response_model=schemas.CorteCajaOut)
def crear_corte_caja(
    datos: schemas.CorteCajaCrear,
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(auth.requiere_permiso("caja.cortar")),
):
    dia = datos.fecha_corte or date.today()
    totales = _totales_ventas_dia(db, dia)
    diferencia = round((datos.efectivo_contado or 0) - totales["total_efectivo"], 2)
    nuevo = models.CorteCaja(
        fecha_corte=datetime.combine(dia, datetime.min.time()),
        usuario_id=usuario.id,
        num_ventas=totales["num_ventas"],
        total_ventas=totales["total_ventas"],
        total_efectivo=totales["total_efectivo"],
        total_tarjeta=totales["total_tarjeta"],
        total_transferencia=totales["total_transferencia"],
        efectivo_contado=datos.efectivo_contado,
        diferencia=diferencia,
        notas=datos.notas or "",
    )
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo


@app.get("/api/caja/cortes", response_model=List[schemas.CorteCajaOut])
def listar_cortes_caja(
    limite: int = 30,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("caja.ver", "caja.cortar")),
):
    return (
        db.query(models.CorteCaja)
        .order_by(models.CorteCaja.fecha_registro.desc())
        .limit(max(1, min(limite, 100)))
        .all()
    )


# ============================================================
#  REPORTES
# ============================================================

@app.get("/api/reportes/resumen-dia")
def resumen_del_dia(
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(auth.obtener_usuario_actual),
):
    """
    Foto del día para el dueño/cajero: ventas, métodos de pago, top 3
    productos y alertas de stock. Cualquier usuario autenticado puede verlo
    (es el tablero del día, no un reporte sensible de periodos largos).
    """
    hoy = date.today()
    totales = _totales_ventas_dia(db, hoy)
    inicio = datetime.combine(hoy, datetime.min.time())
    fin = inicio + timedelta(days=1)

    top = (
        db.query(
            models.Producto.nombre,
            func.sum(models.VentaDetalle.cantidad).label("unidades"),
            func.sum(models.VentaDetalle.subtotal).label("ingresos"),
        )
        .join(models.VentaDetalle, models.VentaDetalle.producto_id == models.Producto.id)
        .join(models.Venta, models.Venta.id == models.VentaDetalle.venta_id)
        .filter(
            models.Venta.cancelada == False,
            models.Venta.fecha >= inicio,
            models.Venta.fecha < fin,
        )
        .group_by(models.Producto.id)
        .order_by(func.sum(models.VentaDetalle.cantidad).desc())
        .limit(3)
        .all()
    )

    bajo_stock = (
        db.query(models.Producto)
        .filter(
            models.Producto.activo == True,
            models.Producto.stock <= models.Producto.stock_minimo,
        )
        .count()
    )

    ya_hay_corte = (
        db.query(models.CorteCaja)
        .filter(models.CorteCaja.fecha_corte >= inicio, models.CorteCaja.fecha_corte < fin)
        .count()
        > 0
    )

    ticket = (
        round(totales["total_ventas"] / totales["num_ventas"], 2)
        if totales["num_ventas"]
        else 0
    )

    return {
        "fecha": hoy.isoformat(),
        "num_ventas": totales["num_ventas"],
        "total_ingresos": totales["total_ventas"],
        "ticket_promedio": ticket,
        "efectivo": totales["total_efectivo"],
        "tarjeta": totales["total_tarjeta"],
        "transferencia": totales["total_transferencia"],
        "top_productos": [
            {
                "nombre": r.nombre,
                "unidades": float(r.unidades or 0),
                "ingresos": round(float(r.ingresos or 0), 2),
            }
            for r in top
        ],
        "productos_bajo_stock": bajo_stock,
        "ya_hay_corte": ya_hay_corte,
        "usuario": usuario.nombre_completo or usuario.username,
    }


@app.get("/api/reportes/gastos")
def resumen_gastos(
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("reportes.ver")),
):
    """
    Gasto en compras de mercancía: suma de (cantidad * costo_unitario) de
    los movimientos de tipo 'entrada' que traen costo capturado. Las
    entradas sin costo (ajustes, devoluciones de venta cancelada, etc.)
    no cuentan como gasto porque no representan dinero que hayas pagado.
    """
    query = (
        db.query(models.MovimientoInventario, models.Producto)
        .join(models.Producto, models.Producto.id == models.MovimientoInventario.producto_id)
        .filter(
            models.MovimientoInventario.tipo == "entrada",
            models.MovimientoInventario.costo_unitario.isnot(None),
        )
    )
    if fecha_inicio:
        query = query.filter(models.MovimientoInventario.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.MovimientoInventario.fecha < fecha_fin + timedelta(days=1))

    filas = query.order_by(models.MovimientoInventario.fecha.desc()).all()

    total_gastado = sum(m.cantidad * m.costo_unitario for m, _ in filas)
    detalle = [
        {
            "fecha": m.fecha,
            "producto": p.nombre,
            "cantidad": m.cantidad,
            "costo_unitario": round(m.costo_unitario, 2),
            "subtotal": round(m.cantidad * m.costo_unitario, 2),
            "motivo": m.motivo or "",
        }
        for m, p in filas
    ]

    return {
        "num_compras": len(filas),
        "total_gastado": round(total_gastado, 2),
        "detalle": detalle,
    }


@app.get("/api/reportes/mermas")
def resumen_mermas(
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("reportes.ver")),
):
    """
    Dinero perdido en mermas: suma de (cantidad * costo_unitario) de los
    movimientos de tipo 'salida'. El costo_unitario de una salida es el
    costo que tenía el producto en ese momento (no representa una compra,
    solo sirve para valorizar la pérdida). Las salidas de un producto sin
    costo registrado no cuentan porque no se puede valorizar la pérdida.
    """
    query = (
        db.query(models.MovimientoInventario, models.Producto)
        .join(models.Producto, models.Producto.id == models.MovimientoInventario.producto_id)
        .filter(
            models.MovimientoInventario.tipo == "salida",
            models.MovimientoInventario.costo_unitario.isnot(None),
        )
    )
    if fecha_inicio:
        query = query.filter(models.MovimientoInventario.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.MovimientoInventario.fecha < fecha_fin + timedelta(days=1))

    filas = query.order_by(models.MovimientoInventario.fecha.desc()).all()

    total_perdido = sum(m.cantidad * m.costo_unitario for m, _ in filas)
    detalle = [
        {
            "fecha": m.fecha,
            "producto": p.nombre,
            "cantidad": m.cantidad,
            "costo_unitario": round(m.costo_unitario, 2),
            "subtotal": round(m.cantidad * m.costo_unitario, 2),
            "motivo": m.motivo or "",
        }
        for m, p in filas
    ]

    return {
        "num_mermas": len(filas),
        "total_perdido": round(total_perdido, 2),
        "detalle": detalle,
    }


@app.get("/api/reportes/margen-bruto")
def resumen_margen_bruto(
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("reportes.ver")),
):
    """
    Ganancia bruta real de lo que se vendió: (precio de venta - costo del
    producto) * cantidad, sumado sobre los detalles de ventas no
    canceladas del periodo.

    A diferencia de la "utilidad neta" (que resta lo GASTADO en compras
    ese mismo periodo, sin importar cuándo se vendió esa mercancía), este
    número no se distorsiona si compraste mucho stock de golpe: refleja
    el margen de lo que efectivamente saliste a vender.

    Limitación: usa el costo ACTUAL registrado en el producto, no el costo
    exacto que tenía el día de cada venta (el sistema no guarda ese
    historial). Si el costo de un producto cambió mucho desde entonces,
    el margen de ventas viejas es una aproximación, no un número exacto.
    """
    query = (
        db.query(models.VentaDetalle, models.Producto)
        .join(models.Producto, models.Producto.id == models.VentaDetalle.producto_id)
        .join(models.Venta, models.Venta.id == models.VentaDetalle.venta_id)
        .filter(models.Venta.cancelada == False)
    )
    if fecha_inicio:
        query = query.filter(models.Venta.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.Venta.fecha < fecha_fin + timedelta(days=1))

    total_ingresos = 0.0
    total_costo = 0.0
    for detalle, producto in query.all():
        total_ingresos += detalle.subtotal
        total_costo += (producto.costo or 0) * detalle.cantidad

    margen_bruto = total_ingresos - total_costo
    porcentaje = (margen_bruto / total_ingresos * 100) if total_ingresos else 0

    return {
        "total_ingresos": round(total_ingresos, 2),
        "total_costo_vendido": round(total_costo, 2),
        "margen_bruto": round(margen_bruto, 2),
        "porcentaje_margen": round(porcentaje, 1),
    }


@app.get("/api/reportes/resumen")
def resumen_ventas(
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("reportes.ver")),
):
    query = db.query(models.Venta).filter(models.Venta.cancelada == False)
    if fecha_inicio:
        query = query.filter(models.Venta.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.Venta.fecha < fecha_fin + timedelta(days=1))

    ventas = query.all()
    total_ingresos = sum(v.total for v in ventas)
    num_ventas = len(ventas)

    return {
        "num_ventas": num_ventas,
        "total_ingresos": round(total_ingresos, 2),
        "ticket_promedio": round(total_ingresos / num_ventas, 2) if num_ventas else 0,
    }


@app.get("/api/reportes/top-productos")
def top_productos(
    limite: int = 10,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("reportes.ver")),
):
    query = (
        db.query(
            models.Producto.nombre,
            func.sum(models.VentaDetalle.cantidad).label("unidades_vendidas"),
            func.sum(models.VentaDetalle.subtotal).label("ingresos"),
        )
        .join(models.VentaDetalle, models.VentaDetalle.producto_id == models.Producto.id)
        .join(models.Venta, models.Venta.id == models.VentaDetalle.venta_id)
        .filter(models.Venta.cancelada == False)
    )
    if fecha_inicio:
        query = query.filter(models.Venta.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.Venta.fecha < fecha_fin + timedelta(days=1))

    resultados = (
        query.group_by(models.Producto.id)
        .order_by(func.sum(models.VentaDetalle.cantidad).desc())
        .limit(limite)
        .all()
    )
    return [
        {"nombre": r.nombre, "unidades_vendidas": int(r.unidades_vendidas), "ingresos": round(r.ingresos, 2)}
        for r in resultados
    ]


@app.get("/api/reportes/ventas-por-dia")
def ventas_por_dia(
    dias: int = 7,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("reportes.ver")),
):
    desde = datetime.now() - timedelta(days=dias)
    ventas = db.query(models.Venta).filter(
        models.Venta.fecha >= desde, models.Venta.cancelada == False
    ).all()

    resumen = {}
    for v in ventas:
        clave = v.fecha.strftime("%Y-%m-%d")
        resumen[clave] = resumen.get(clave, 0) + v.total

    return [{"fecha": k, "total": round(v, 2)} for k, v in sorted(resumen.items())]


def _rango_mes_actual_parcial():
    """Del día 1 del mes actual a hoy."""
    hoy = date.today()
    return hoy.replace(day=1), hoy


def _rango_mes_anterior_parcial():
    """Del día 1 del mes anterior al mismo número de día que lleva el mes
    actual (o al último día del mes anterior si este es más corto), para
    que la comparación sea entre periodos de la misma duración y no entre
    un mes completo contra uno que apenas va a la mitad."""
    hoy = date.today()
    primer_dia_mes_actual = hoy.replace(day=1)
    ultimo_dia_mes_anterior = primer_dia_mes_actual - timedelta(days=1)
    inicio_mes_anterior = ultimo_dia_mes_anterior.replace(day=1)
    dia_equivalente = min(hoy.day, ultimo_dia_mes_anterior.day)
    fin_mes_anterior = inicio_mes_anterior.replace(day=dia_equivalente)
    return inicio_mes_anterior, fin_mes_anterior


def _metricas_periodo(db: Session, fecha_inicio: Optional[date], fecha_fin: Optional[date]):
    """Calcula, para un rango de fechas, las métricas clave del negocio:
    ventas, gasto en compras de mercancía, pérdida por mermas, otros gastos
    fijos y la utilidad neta resultante. Se usa tanto para el reporte del
    periodo elegido como para la comparativa mes anterior vs mes actual."""

    q_ventas = db.query(models.Venta).filter(models.Venta.cancelada == False)
    if fecha_inicio:
        q_ventas = q_ventas.filter(models.Venta.fecha >= fecha_inicio)
    if fecha_fin:
        q_ventas = q_ventas.filter(models.Venta.fecha < fecha_fin + timedelta(days=1))
    ventas = q_ventas.all()
    total_ventas = sum(v.total or 0 for v in ventas)

    q_compras = db.query(models.MovimientoInventario).filter(
        models.MovimientoInventario.tipo == "entrada",
        models.MovimientoInventario.costo_unitario.isnot(None),
    )
    if fecha_inicio:
        q_compras = q_compras.filter(models.MovimientoInventario.fecha >= fecha_inicio)
    if fecha_fin:
        q_compras = q_compras.filter(models.MovimientoInventario.fecha < fecha_fin + timedelta(days=1))
    compras = q_compras.all()
    total_compras = sum((m.cantidad or 0) * (m.costo_unitario or 0) for m in compras)

    q_mermas = db.query(models.MovimientoInventario).filter(
        models.MovimientoInventario.tipo == "salida",
        models.MovimientoInventario.costo_unitario.isnot(None),
    )
    if fecha_inicio:
        q_mermas = q_mermas.filter(models.MovimientoInventario.fecha >= fecha_inicio)
    if fecha_fin:
        q_mermas = q_mermas.filter(models.MovimientoInventario.fecha < fecha_fin + timedelta(days=1))
    mermas = q_mermas.all()
    total_mermas = sum((m.cantidad or 0) * (m.costo_unitario or 0) for m in mermas)

    q_gastos = db.query(models.OtroGasto)
    if fecha_inicio:
        q_gastos = q_gastos.filter(models.OtroGasto.fecha >= fecha_inicio)
    if fecha_fin:
        q_gastos = q_gastos.filter(models.OtroGasto.fecha < fecha_fin + timedelta(days=1))
    otros_gastos = q_gastos.all()
    total_otros = sum(g.monto or 0 for g in otros_gastos)

    utilidad_neta = total_ventas - total_compras - total_mermas - total_otros

    return {
        "num_ventas": len(ventas), "total_ventas": round(total_ventas, 2),
        "num_compras": len(compras), "total_compras": round(total_compras, 2),
        "num_mermas": len(mermas), "total_mermas": round(total_mermas, 2),
        "num_otros_gastos": len(otros_gastos), "total_otros_gastos": round(total_otros, 2),
        "utilidad_neta": round(utilidad_neta, 2),
    }


@app.get("/api/reportes/comparativa-mensual")
def comparativa_mensual(
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("reportes.ver")),
):
    """
    Compara ventas, gastos en compras, mermas y otros gastos del mes en
    curso contra el mismo número de días del mes anterior (para que la
    comparación sea justa aunque el mes actual todavía no termine).
    """
    inicio_actual, fin_actual = _rango_mes_actual_parcial()
    inicio_anterior, fin_anterior = _rango_mes_anterior_parcial()

    actual = _metricas_periodo(db, inicio_actual, fin_actual)
    anterior = _metricas_periodo(db, inicio_anterior, fin_anterior)

    return {
        "mes_actual": {"inicio": inicio_actual.isoformat(), "fin": fin_actual.isoformat(), **actual},
        "mes_anterior": {"inicio": inicio_anterior.isoformat(), "fin": fin_anterior.isoformat(), **anterior},
    }


def _obtener_datos_reporte(db: Session, fecha_inicio: Optional[date], fecha_fin: Optional[date]):
    """Junta todo lo necesario para los reportes exportables: resumen y
    detalle de ventas, top de productos, gastos en compras de mercancía,
    mermas y pérdidas, otros gastos fijos, la utilidad neta del periodo, y
    la comparativa mes anterior vs mes actual (esta última siempre se
    calcula sobre el mes en curso, sin importar el rango de fechas elegido
    para el resto del reporte, porque su propósito es dar una foto fija de
    cómo va el negocio ahora mismo)."""
    query = db.query(models.Venta).filter(models.Venta.cancelada == False)
    if fecha_inicio:
        query = query.filter(models.Venta.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.Venta.fecha < fecha_fin + timedelta(days=1))
    ventas = query.order_by(models.Venta.fecha.desc()).all()

    total_ingresos = sum(v.total or 0 for v in ventas)
    num_ventas = len(ventas)
    ticket_promedio = total_ingresos / num_ventas if num_ventas else 0

    top_query = (
        db.query(
            models.Producto.nombre,
            func.sum(models.VentaDetalle.cantidad).label("unidades"),
            func.sum(models.VentaDetalle.subtotal).label("ingresos"),
        )
        .join(models.VentaDetalle, models.VentaDetalle.producto_id == models.Producto.id)
        .join(models.Venta, models.Venta.id == models.VentaDetalle.venta_id)
        .filter(models.Venta.cancelada == False)
    )
    if fecha_inicio:
        top_query = top_query.filter(models.Venta.fecha >= fecha_inicio)
    if fecha_fin:
        top_query = top_query.filter(models.Venta.fecha < fecha_fin + timedelta(days=1))
    top = (
        top_query.group_by(models.Producto.id)
        .order_by(func.sum(models.VentaDetalle.cantidad).desc())
        .limit(15)
        .all()
    )

    # --- Gastos en compras de mercancía (entradas de inventario con costo) ---
    q_compras = (
        db.query(models.MovimientoInventario, models.Producto)
        .join(models.Producto, models.Producto.id == models.MovimientoInventario.producto_id)
        .filter(
            models.MovimientoInventario.tipo == "entrada",
            models.MovimientoInventario.costo_unitario.isnot(None),
        )
    )
    if fecha_inicio:
        q_compras = q_compras.filter(models.MovimientoInventario.fecha >= fecha_inicio)
    if fecha_fin:
        q_compras = q_compras.filter(models.MovimientoInventario.fecha < fecha_fin + timedelta(days=1))
    filas_compras = q_compras.order_by(models.MovimientoInventario.fecha.desc()).all()
    total_compras = sum((m.cantidad or 0) * (m.costo_unitario or 0) for m, _ in filas_compras)

    # --- Mermas y pérdidas (salidas de inventario con costo) ---
    q_mermas = (
        db.query(models.MovimientoInventario, models.Producto)
        .join(models.Producto, models.Producto.id == models.MovimientoInventario.producto_id)
        .filter(
            models.MovimientoInventario.tipo == "salida",
            models.MovimientoInventario.costo_unitario.isnot(None),
        )
    )
    if fecha_inicio:
        q_mermas = q_mermas.filter(models.MovimientoInventario.fecha >= fecha_inicio)
    if fecha_fin:
        q_mermas = q_mermas.filter(models.MovimientoInventario.fecha < fecha_fin + timedelta(days=1))
    filas_mermas = q_mermas.order_by(models.MovimientoInventario.fecha.desc()).all()
    total_mermas = sum((m.cantidad or 0) * (m.costo_unitario or 0) for m, _ in filas_mermas)

    # --- Otros gastos (renta, luz, sueldos, etc.) ---
    q_otros = db.query(models.OtroGasto)
    if fecha_inicio:
        q_otros = q_otros.filter(models.OtroGasto.fecha >= fecha_inicio)
    if fecha_fin:
        q_otros = q_otros.filter(models.OtroGasto.fecha < fecha_fin + timedelta(days=1))
    otros_gastos = q_otros.order_by(models.OtroGasto.fecha.desc()).all()
    total_otros = sum(g.monto or 0 for g in otros_gastos)

    utilidad_neta = total_ingresos - total_compras - total_mermas - total_otros

    # --- Comparativa mes anterior vs mes actual ---
    inicio_actual, fin_actual = _rango_mes_actual_parcial()
    inicio_anterior, fin_anterior = _rango_mes_anterior_parcial()
    comparativa = {
        "mes_actual": {
            "inicio": inicio_actual.isoformat(), "fin": fin_actual.isoformat(),
            **_metricas_periodo(db, inicio_actual, fin_actual),
        },
        "mes_anterior": {
            "inicio": inicio_anterior.isoformat(), "fin": fin_anterior.isoformat(),
            **_metricas_periodo(db, inicio_anterior, fin_anterior),
        },
    }

    return {
        "resumen": {"num_ventas": num_ventas, "total_ingresos": total_ingresos, "ticket_promedio": ticket_promedio},
        "top": [{"nombre": r.nombre, "unidades": int(r.unidades), "ingresos": r.ingresos} for r in top],
        "ventas": ventas,
        "compras": {
            "total": round(total_compras, 2),
            "detalle": [
                {
                    "fecha": m.fecha, "producto": p.nombre, "cantidad": m.cantidad,
                    "costo_unitario": round(m.costo_unitario, 2),
                    "subtotal": round((m.cantidad or 0) * (m.costo_unitario or 0), 2),
                    "motivo": m.motivo or "",
                }
                for m, p in filas_compras
            ],
        },
        "mermas": {
            "total": round(total_mermas, 2),
            "detalle": [
                {
                    "fecha": m.fecha, "producto": p.nombre, "cantidad": m.cantidad,
                    "costo_unitario": round(m.costo_unitario, 2),
                    "subtotal": round((m.cantidad or 0) * (m.costo_unitario or 0), 2),
                    "motivo": m.motivo or "",
                }
                for m, p in filas_mermas
            ],
        },
        "otros_gastos": {
            "total": round(total_otros, 2),
            "detalle": [
                {
                    "fecha": g.fecha, "concepto": g.concepto, "categoria": g.categoria or "otro",
                    "monto": round(g.monto or 0, 2), "notas": g.notas or "",
                }
                for g in otros_gastos
            ],
        },
        "utilidad_neta": round(utilidad_neta, 2),
        "comparativa": comparativa,
    }


@app.get("/api/reportes/exportar/excel")
def exportar_excel(
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("reportes.ver")),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    datos = _obtener_datos_reporte(db, fecha_inicio, fecha_fin)
    nombre_tienda = _nombre_tienda(db)
    wb = Workbook()
    encabezado_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    encabezado_font = Font(color="FFFFFF", bold=True)

    # --- Hoja Resumen ---
    ws = wb.active
    ws.title = "Resumen"
    ws.append([nombre_tienda])
    ws.append(["Reporte de ventas"])
    rango = f"{fecha_inicio or 'inicio'} a {fecha_fin or 'hoy'}"
    ws.append([f"Periodo: {rango}"])
    ws.append([])
    ws.append(["Número de ventas", datos["resumen"]["num_ventas"]])
    ws.append(["Total de ingresos", round(datos["resumen"]["total_ingresos"], 2)])
    ws.append(["Ticket promedio", round(datos["resumen"]["ticket_promedio"], 2)])
    ws.append([])
    ws.append(["Gastos en compras de mercancía", datos["compras"]["total"]])
    ws.append(["Pérdida por mermas", datos["mermas"]["total"]])
    ws.append(["Otros gastos (renta, luz, sueldos, etc.)", datos["otros_gastos"]["total"]])
    ws.append(["Utilidad neta del periodo", datos["utilidad_neta"]])

    # --- Hoja Top productos ---
    ws2 = wb.create_sheet("Top productos")
    ws2.append(["Producto", "Unidades vendidas", "Ingresos"])
    for c in ws2[1]:
        c.font = encabezado_font
        c.fill = encabezado_fill
    for p in datos["top"]:
        ws2.append([p["nombre"], p["unidades"], round(p["ingresos"], 2)])

    # --- Hoja Detalle de ventas ---
    ws3 = wb.create_sheet("Ventas")
    ws3.append(["ID", "Fecha", "Total", "Método de pago", "Vendedor"])
    for c in ws3[1]:
        c.font = encabezado_font
        c.fill = encabezado_fill
    for v in datos["ventas"]:
        vendedor = (v.usuario.nombre_completo or v.usuario.username) if v.usuario else "—"
        total_v = v.total if v.total is not None else 0
        fecha_v = v.fecha.strftime("%Y-%m-%d %H:%M") if v.fecha else "—"
        ws3.append([v.id, fecha_v, round(total_v, 2), v.metodo_pago, vendedor])

    # --- Hoja Compras de mercancía ---
    ws4 = wb.create_sheet("Compras")
    ws4.append(["Fecha", "Producto", "Cantidad", "Costo unitario", "Subtotal", "Motivo"])
    for c in ws4[1]:
        c.font = encabezado_font
        c.fill = encabezado_fill
    for cpr in datos["compras"]["detalle"]:
        fecha_c = cpr["fecha"].strftime("%Y-%m-%d %H:%M") if cpr["fecha"] else "—"
        ws4.append([fecha_c, cpr["producto"], cpr["cantidad"], cpr["costo_unitario"], cpr["subtotal"], cpr["motivo"]])
    ws4.append([])
    ws4.append(["", "", "", "Total gastado", datos["compras"]["total"], ""])

    # --- Hoja Mermas y pérdidas ---
    ws5 = wb.create_sheet("Mermas")
    ws5.append(["Fecha", "Producto", "Cantidad", "Costo unitario", "Subtotal", "Motivo"])
    for c in ws5[1]:
        c.font = encabezado_font
        c.fill = encabezado_fill
    for m in datos["mermas"]["detalle"]:
        fecha_m = m["fecha"].strftime("%Y-%m-%d %H:%M") if m["fecha"] else "—"
        ws5.append([fecha_m, m["producto"], m["cantidad"], m["costo_unitario"], m["subtotal"], m["motivo"]])
    ws5.append([])
    ws5.append(["", "", "", "Total perdido", datos["mermas"]["total"], ""])

    # --- Hoja Otros gastos ---
    ws6 = wb.create_sheet("Otros gastos")
    ws6.append(["Fecha", "Concepto", "Categoría", "Monto", "Notas"])
    for c in ws6[1]:
        c.font = encabezado_font
        c.fill = encabezado_fill
    for g in datos["otros_gastos"]["detalle"]:
        fecha_g = g["fecha"].strftime("%Y-%m-%d %H:%M") if g["fecha"] else "—"
        ws6.append([fecha_g, g["concepto"], g["categoria"], g["monto"], g["notas"]])
    ws6.append([])
    ws6.append(["", "", "Total", datos["otros_gastos"]["total"], ""])

    # --- Hoja Comparativo mensual ---
    ws7 = wb.create_sheet("Comparativo mensual")
    comp = datos["comparativa"]
    ma, mp = comp["mes_actual"], comp["mes_anterior"]
    ws7.append(["Comparativa: mes anterior vs mes actual"])
    ws7.append([f"Mes anterior: {mp['inicio']} a {mp['fin']}   |   Mes actual: {ma['inicio']} a {ma['fin']}"])
    ws7.append([])
    ws7.append(["Métrica", "Mes anterior", "Mes actual", "Variación", "Variación %"])
    for c in ws7[4]:
        c.font = encabezado_font
        c.fill = encabezado_fill

    def _fila_comparativa(nombre, anterior, actual):
        variacion = actual - anterior
        variacion_pct = (variacion / anterior * 100) if anterior else (100.0 if actual else 0.0)
        ws7.append([nombre, round(anterior, 2), round(actual, 2), round(variacion, 2), f"{variacion_pct:.1f}%"])

    _fila_comparativa("Ventas", mp["total_ventas"], ma["total_ventas"])
    _fila_comparativa("Gastos en compras de mercancía", mp["total_compras"], ma["total_compras"])
    _fila_comparativa("Pérdida por mermas", mp["total_mermas"], ma["total_mermas"])
    _fila_comparativa("Otros gastos (renta, luz, sueldos, etc.)", mp["total_otros_gastos"], ma["total_otros_gastos"])
    _fila_comparativa("Utilidad neta", mp["utilidad_neta"], ma["utilidad_neta"])

    for hoja in (ws, ws2, ws3, ws4, ws5, ws6, ws7):
        for columna in hoja.columns:
            ancho = max((len(str(c.value)) for c in columna if c.value is not None), default=10)
            hoja.column_dimensions[columna[0].column_letter].width = min(ancho + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_ventas.xlsx"},
    )


def _escapar_pdf(texto) -> str:
    """Escapa caracteres especiales para reportlab Paragraph (XML)."""
    if texto is None:
        return ""
    return (
        str(texto)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _moneda(valor) -> str:
    """Formatea un número como moneda, sin romperse si viene None o texto.
    Los negativos se muestran como -$123.45 (no $-123.45)."""
    try:
        numero = float(valor or 0)
    except (TypeError, ValueError):
        return "$0.00"
    signo = "-" if numero < 0 else ""
    return f"{signo}${abs(numero):,.2f}"


def _num(valor) -> int:
    try:
        return int(valor or 0)
    except (TypeError, ValueError):
        return 0


def _texto(valor, alterno="—") -> str:
    if valor is None:
        return alterno
    return str(valor)


@app.get("/api/reportes/exportar/pdf")
def exportar_pdf(
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    db: Session = Depends(get_db),
    _usuario: models.Usuario = Depends(auth.requiere_permiso("reportes.ver")),
):
    """
    Genera el reporte de ventas en PDF con un diseño profesional (tablas,
    encabezados de sección y numeración de página) usando reportlab/platypus.

    Todo el proceso está envuelto en try/except: si algo falla, se devuelve
    un error HTTP 500 con el detalle real en vez de dejar que la excepción
    se pierda y el cliente termine descargando un archivo de 0 bytes.
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    )
    from reportlab.pdfgen import canvas as pdf_canvas

    try:
        datos = _obtener_datos_reporte(db, fecha_inicio, fecha_fin)
        nombre_tienda = _nombre_tienda(db)
        r = datos["resumen"]
        top = datos["top"]
        ventas = datos["ventas"]

        AZUL = colors.HexColor("#2563EB")
        AZUL_CLARO = colors.HexColor("#EFF4FF")
        GRIS_CLARO = colors.HexColor("#F5F5F7")
        GRIS_TEXTO = colors.HexColor("#374151")

        estilos = getSampleStyleSheet()
        titulo_style = ParagraphStyle(
            "TituloReporte", parent=estilos["Title"], fontSize=20,
            textColor=AZUL, spaceAfter=2,
        )
        subtitulo_style = ParagraphStyle(
            "Subtitulo", parent=estilos["Normal"], fontSize=10,
            textColor=GRIS_TEXTO, spaceAfter=2,
        )
        seccion_style = ParagraphStyle(
            "Seccion", parent=estilos["Heading2"], fontSize=13,
            textColor=colors.white, spaceAfter=0, spaceBefore=0,
            leading=16,
        )
        celda_style = ParagraphStyle(
            "Celda", parent=estilos["Normal"], fontSize=9, leading=11,
        )
        celda_r_style = ParagraphStyle(
            "CeldaR", parent=celda_style, alignment=TA_RIGHT,
        )
        vacio_style = ParagraphStyle(
            "Vacio", parent=estilos["Normal"], fontSize=10,
            textColor=colors.grey, alignment=TA_CENTER,
        )

        def encabezado_seccion(texto):
            """Barra de color con el título de cada sección, estilo Excel."""
            t = Table([[Paragraph(texto, seccion_style)]], colWidths=[17 * cm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), AZUL),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            return t

        elementos = []

        # --- Encabezado del reporte ---
        elementos.append(Paragraph(_escapar_pdf(nombre_tienda), titulo_style))
        elementos.append(Paragraph("Reporte de ventas", subtitulo_style))
        rango = f"Periodo: {fecha_inicio or 'inicio'} al {fecha_fin or 'día de hoy'}"
        elementos.append(Paragraph(rango, subtitulo_style))
        elementos.append(Paragraph(
            f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}",
            subtitulo_style,
        ))
        elementos.append(Spacer(1, 0.5 * cm))

        # --- Resumen ---
        utilidad_neta = float(datos.get("utilidad_neta") or 0)
        color_utilidad = colors.HexColor("#059669") if utilidad_neta >= 0 else colors.HexColor("#DC2626")
        celda_utilidad_style = ParagraphStyle(
            "CeldaUtilidad", parent=celda_style, fontName="Helvetica-Bold",
        )
        celda_utilidad_r_style = ParagraphStyle(
            "CeldaUtilidadR", parent=celda_r_style, fontName="Helvetica-Bold", textColor=color_utilidad,
        )

        elementos.append(encabezado_seccion("Resumen general"))
        tabla_resumen = Table(
            [
                [Paragraph("Número de ventas", celda_style), Paragraph(str(_num(r["num_ventas"])), celda_r_style)],
                [Paragraph("Total de ingresos por ventas", celda_style), Paragraph(_moneda(r["total_ingresos"]), celda_r_style)],
                [Paragraph("Ticket promedio", celda_style), Paragraph(_moneda(r["ticket_promedio"]), celda_r_style)],
                [Paragraph("Gastos en compras de mercancía", celda_style), Paragraph(_moneda(datos["compras"]["total"]), celda_r_style)],
                [Paragraph("Pérdida por mermas", celda_style), Paragraph(_moneda(datos["mermas"]["total"]), celda_r_style)],
                [Paragraph("Otros gastos (renta, luz, sueldos, etc.)", celda_style), Paragraph(_moneda(datos["otros_gastos"]["total"]), celda_r_style)],
                [Paragraph("Utilidad neta del periodo", celda_utilidad_style), Paragraph(_moneda(utilidad_neta), celda_utilidad_r_style)],
            ],
            colWidths=[12 * cm, 5 * cm],
        )
        tabla_resumen.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), AZUL_CLARO),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [AZUL_CLARO, colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LINEABOVE", (0, -1), (-1, -1), 1, AZUL),
        ]))
        elementos.append(tabla_resumen)
        elementos.append(Paragraph(
            "Utilidad neta = ingresos por ventas − gastos en compras de mercancía − pérdida por mermas − otros gastos.",
            vacio_style,
        ))
        elementos.append(Spacer(1, 0.6 * cm))

        # --- Top productos ---
        elementos.append(encabezado_seccion("Top productos vendidos"))
        if top:
            total_ingresos = float(r["total_ingresos"] or 0)
            filas = [[
                Paragraph("#", celda_style), Paragraph("Producto", celda_style),
                Paragraph("Unidades", celda_r_style), Paragraph("Ingresos", celda_r_style),
                Paragraph("% del total", celda_r_style),
            ]]
            for i, p in enumerate(top[:15], start=1):
                ingresos_p = float(p.get("ingresos") or 0)
                porcentaje = (ingresos_p / total_ingresos * 100) if total_ingresos else 0
                filas.append([
                    Paragraph(str(i), celda_style),
                    Paragraph(_texto(p.get("nombre"), "Producto sin nombre"), celda_style),
                    Paragraph(str(_num(p.get("unidades"))), celda_r_style),
                    Paragraph(_moneda(ingresos_p), celda_r_style),
                    Paragraph(f"{porcentaje:.1f}%", celda_r_style),
                ])
            tabla_top = Table(filas, colWidths=[1 * cm, 8 * cm, 2.7 * cm, 3 * cm, 2.3 * cm], repeatRows=1)
            tabla_top.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), AZUL),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_CLARO]),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]))
            elementos.append(tabla_top)
        else:
            elementos.append(Paragraph(
                "No se vendieron productos en el periodo seleccionado.", vacio_style,
            ))
        elementos.append(Spacer(1, 0.6 * cm))

        # --- Historial de ventas ---
        elementos.append(encabezado_seccion(f"Historial de ventas ({len(ventas)} en total)"))
        LIMITE_DETALLE = 300
        if ventas:
            filas_v = [[
                Paragraph("ID", celda_style), Paragraph("Fecha", celda_style),
                Paragraph("Total", celda_r_style), Paragraph("Método de pago", celda_style),
                Paragraph("Vendedor", celda_style),
            ]]
            for v in ventas[:LIMITE_DETALLE]:
                if v.usuario:
                    vendedor = _texto(v.usuario.nombre_completo or v.usuario.username)
                else:
                    vendedor = "—"
                fecha_txt = v.fecha.strftime("%d/%m/%Y %H:%M") if v.fecha else "—"
                filas_v.append([
                    Paragraph(f"#{_num(v.id)}", celda_style),
                    Paragraph(fecha_txt, celda_style),
                    Paragraph(_moneda(v.total), celda_r_style),
                    Paragraph(_texto(v.metodo_pago), celda_style),
                    Paragraph(vendedor, celda_style),
                ])
            tabla_ventas = Table(
                filas_v, colWidths=[1.8 * cm, 3.5 * cm, 3 * cm, 4 * cm, 4.7 * cm], repeatRows=1,
            )
            tabla_ventas.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), AZUL),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_CLARO]),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]))
            elementos.append(tabla_ventas)
            if len(ventas) > LIMITE_DETALLE:
                elementos.append(Spacer(1, 0.3 * cm))
                elementos.append(Paragraph(
                    f"Mostrando las primeras {LIMITE_DETALLE} de {len(ventas)} ventas. "
                    "Usa la exportación a Excel para ver el listado completo.",
                    vacio_style,
                ))
        else:
            elementos.append(Paragraph(
                "No hay ventas registradas en el periodo seleccionado.", vacio_style,
            ))

        # --- Gastos en compras de mercancía ---
        elementos.append(PageBreak())
        compras_detalle = datos["compras"]["detalle"]
        elementos.append(encabezado_seccion(f"Gastos en compras de mercancía ({len(compras_detalle)} compras)"))
        if compras_detalle:
            filas_c = [[
                Paragraph("Fecha", celda_style), Paragraph("Producto", celda_style),
                Paragraph("Cantidad", celda_r_style), Paragraph("Costo unitario", celda_r_style),
                Paragraph("Subtotal", celda_r_style), Paragraph("Motivo", celda_style),
            ]]
            for cpr in compras_detalle[:LIMITE_DETALLE]:
                fecha_txt = cpr["fecha"].strftime("%d/%m/%Y %H:%M") if cpr["fecha"] else "—"
                filas_c.append([
                    Paragraph(fecha_txt, celda_style),
                    Paragraph(_texto(cpr["producto"], "Producto sin nombre"), celda_style),
                    Paragraph(str(_num(cpr["cantidad"])), celda_r_style),
                    Paragraph(_moneda(cpr["costo_unitario"]), celda_r_style),
                    Paragraph(_moneda(cpr["subtotal"]), celda_r_style),
                    Paragraph(_texto(cpr["motivo"], ""), celda_style),
                ])
            tabla_c = Table(
                filas_c, colWidths=[2.6 * cm, 5.3 * cm, 2 * cm, 2.6 * cm, 2.5 * cm, 2 * cm], repeatRows=1,
            )
            tabla_c.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), AZUL),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_CLARO]),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]))
            elementos.append(tabla_c)
            if len(compras_detalle) > LIMITE_DETALLE:
                elementos.append(Spacer(1, 0.2 * cm))
                elementos.append(Paragraph(
                    f"Mostrando las primeras {LIMITE_DETALLE} de {len(compras_detalle)} compras.", vacio_style,
                ))
            elementos.append(Spacer(1, 0.2 * cm))
            elementos.append(Paragraph(
                f"<b>Total gastado en compras: {_moneda(datos['compras']['total'])}</b>", celda_style,
            ))
        else:
            elementos.append(Paragraph(
                "No se registraron compras de mercancía en el periodo seleccionado.", vacio_style,
            ))
        elementos.append(Spacer(1, 0.6 * cm))

        # --- Mermas y pérdidas ---
        mermas_detalle = datos["mermas"]["detalle"]
        elementos.append(encabezado_seccion(f"Mermas y pérdidas ({len(mermas_detalle)} movimientos)"))
        if mermas_detalle:
            filas_m = [[
                Paragraph("Fecha", celda_style), Paragraph("Producto", celda_style),
                Paragraph("Cantidad", celda_r_style), Paragraph("Costo unitario", celda_r_style),
                Paragraph("Subtotal", celda_r_style), Paragraph("Motivo", celda_style),
            ]]
            for m in mermas_detalle[:LIMITE_DETALLE]:
                fecha_txt = m["fecha"].strftime("%d/%m/%Y %H:%M") if m["fecha"] else "—"
                filas_m.append([
                    Paragraph(fecha_txt, celda_style),
                    Paragraph(_texto(m["producto"], "Producto sin nombre"), celda_style),
                    Paragraph(str(_num(m["cantidad"])), celda_r_style),
                    Paragraph(_moneda(m["costo_unitario"]), celda_r_style),
                    Paragraph(_moneda(m["subtotal"]), celda_r_style),
                    Paragraph(_texto(m["motivo"], ""), celda_style),
                ])
            tabla_m = Table(
                filas_m, colWidths=[2.6 * cm, 5.3 * cm, 2 * cm, 2.6 * cm, 2.5 * cm, 2 * cm], repeatRows=1,
            )
            tabla_m.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DC2626")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_CLARO]),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]))
            elementos.append(tabla_m)
            if len(mermas_detalle) > LIMITE_DETALLE:
                elementos.append(Spacer(1, 0.2 * cm))
                elementos.append(Paragraph(
                    f"Mostrando las primeras {LIMITE_DETALLE} de {len(mermas_detalle)} mermas.", vacio_style,
                ))
            elementos.append(Spacer(1, 0.2 * cm))
            elementos.append(Paragraph(
                f"<b>Total perdido por mermas: {_moneda(datos['mermas']['total'])}</b>", celda_style,
            ))
        else:
            elementos.append(Paragraph(
                "No se registraron mermas ni pérdidas en el periodo seleccionado.", vacio_style,
            ))
        elementos.append(Spacer(1, 0.6 * cm))

        # --- Otros gastos (renta, luz, sueldos, etc.) ---
        otros_detalle = datos["otros_gastos"]["detalle"]
        elementos.append(encabezado_seccion(f"Otros gastos — renta, luz, sueldos, etc. ({len(otros_detalle)})"))
        if otros_detalle:
            filas_g = [[
                Paragraph("Fecha", celda_style), Paragraph("Concepto", celda_style),
                Paragraph("Categoría", celda_style), Paragraph("Monto", celda_r_style),
                Paragraph("Notas", celda_style),
            ]]
            for g in otros_detalle[:LIMITE_DETALLE]:
                fecha_txt = g["fecha"].strftime("%d/%m/%Y %H:%M") if g["fecha"] else "—"
                filas_g.append([
                    Paragraph(fecha_txt, celda_style),
                    Paragraph(_texto(g["concepto"], "Sin concepto"), celda_style),
                    Paragraph(_texto(g["categoria"], "otro").capitalize(), celda_style),
                    Paragraph(_moneda(g["monto"]), celda_r_style),
                    Paragraph(_texto(g["notas"], ""), celda_style),
                ])
            tabla_g = Table(
                filas_g, colWidths=[2.6 * cm, 4.3 * cm, 2.6 * cm, 2.5 * cm, 5 * cm], repeatRows=1,
            )
            tabla_g.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), AZUL),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_CLARO]),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]))
            elementos.append(tabla_g)
            if len(otros_detalle) > LIMITE_DETALLE:
                elementos.append(Spacer(1, 0.2 * cm))
                elementos.append(Paragraph(
                    f"Mostrando los primeros {LIMITE_DETALLE} de {len(otros_detalle)} gastos.", vacio_style,
                ))
            elementos.append(Spacer(1, 0.2 * cm))
            elementos.append(Paragraph(
                f"<b>Total en otros gastos: {_moneda(datos['otros_gastos']['total'])}</b>", celda_style,
            ))
        else:
            elementos.append(Paragraph(
                "No se registraron otros gastos en el periodo seleccionado.", vacio_style,
            ))

        # --- Comparativa: mes anterior vs mes actual ---
        elementos.append(PageBreak())
        comp = datos["comparativa"]
        ma, mp = comp["mes_actual"], comp["mes_anterior"]
        elementos.append(encabezado_seccion("Comparativa: mes anterior vs mes actual"))
        elementos.append(Spacer(1, 0.2 * cm))
        elementos.append(Paragraph(
            f"Mes anterior: {mp['inicio']} al {mp['fin']}  &nbsp;|&nbsp;  "
            f"Mes actual: {ma['inicio']} al {ma['fin']}",
            subtitulo_style,
        ))
        elementos.append(Paragraph(
            "Se compara el mismo número de días transcurridos en ambos meses, "
            "para que la comparación sea justa aunque el mes actual no haya terminado.",
            vacio_style,
        ))
        elementos.append(Spacer(1, 0.3 * cm))

        def _variacion_texto(anterior, actual, favorable_si_sube):
            variacion = actual - anterior
            if anterior:
                pct_texto = f"{(variacion / anterior * 100):+.1f}%"
            else:
                pct_texto = "nuevo" if actual else "0.0%"
            sube = variacion > 0
            if variacion == 0:
                color_hex = "#6B7280"
            elif (sube and favorable_si_sube) or (not sube and not favorable_si_sube):
                color_hex = "#059669"
            else:
                color_hex = "#DC2626"
            signo = "+" if variacion > 0 else ("-" if variacion < 0 else "")
            monto_txt = f"{signo}{_moneda(abs(variacion))}"
            return f'<font color="{color_hex}">{monto_txt} ({pct_texto})</font>'

        filas_comp = [[
            Paragraph("Métrica", celda_style), Paragraph("Mes anterior", celda_r_style),
            Paragraph("Mes actual", celda_r_style), Paragraph("Variación", celda_r_style),
        ]]
        metricas_comp = [
            ("Ventas", "total_ventas", True),
            ("Gastos en compras de mercancía", "total_compras", False),
            ("Pérdida por mermas", "total_mermas", False),
            ("Otros gastos (renta, luz, sueldos, etc.)", "total_otros_gastos", False),
            ("Utilidad neta", "utilidad_neta", True),
        ]
        for nombre, clave, favorable_si_sube in metricas_comp:
            anterior_v = float(mp.get(clave) or 0)
            actual_v = float(ma.get(clave) or 0)
            filas_comp.append([
                Paragraph(nombre, celda_style),
                Paragraph(_moneda(anterior_v), celda_r_style),
                Paragraph(_moneda(actual_v), celda_r_style),
                Paragraph(_variacion_texto(anterior_v, actual_v, favorable_si_sube), celda_r_style),
            ])
        tabla_comp = Table(filas_comp, colWidths=[6.5 * cm, 3.5 * cm, 3.5 * cm, 4 * cm], repeatRows=1)
        tabla_comp.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), AZUL),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_CLARO]),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        elementos.append(tabla_comp)
        elementos.append(Spacer(1, 0.3 * cm))
        elementos.append(Paragraph(
            f"Número de ventas: {_num(mp.get('num_ventas'))} (mes anterior) vs "
            f"{_num(ma.get('num_ventas'))} (mes actual).",
            vacio_style,
        ))

        # --- Pie de página con numeración ---
        def pie_de_pagina(canv: pdf_canvas.Canvas, doc):
            canv.saveState()
            canv.setFont("Helvetica", 8)
            canv.setFillColor(colors.grey)
            canv.drawString(2 * cm, 1.2 * cm, f"{nombre_tienda} — Reporte de ventas")
            canv.drawRightString(
                letter[0] - 2 * cm, 1.2 * cm, f"Página {doc.page}",
            )
            canv.restoreState()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=letter,
            topMargin=2 * cm, bottomMargin=2 * cm,
            leftMargin=2 * cm, rightMargin=2 * cm,
            title=f"{nombre_tienda} — Reporte de ventas",
        )
        doc.build(elementos, onFirstPage=pie_de_pagina, onLaterPages=pie_de_pagina)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=reporte_ventas.pdf"},
        )

    except Exception as e:
        # Si algo falla, se informa el error real en vez de dejar que el
        # cliente reciba una respuesta vacía / un archivo de 0 bytes.
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"No se pudo generar el PDF: {e}")


# ============================================================
#  FRONTEND (archivos estáticos)
# ============================================================

app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
def index():
    return FileResponse("static/index.html")
